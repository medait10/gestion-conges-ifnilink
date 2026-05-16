
import os
import time
from collections import defaultdict, deque
import base64, json, calendar, shutil, glob, secrets, hashlib
from datetime import date, datetime, timedelta
from email.mime.text import MIMEText
from functools import wraps

from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, abort
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from google_auth_oauthlib.flow import Flow
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import stripe

load_dotenv()
os.environ["OAUTHLIB_INSECURE_TRANSPORT"] = "1"
os.environ["OAUTHLIB_RELAX_TOKEN_SCOPE"] = "1"
app = Flask(__name__, instance_relative_config=True)
app.secret_key = os.getenv("SECRET_KEY", "dev-secret-change-me")
ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "aitelmalemmohamed@gmail.com").lower()
stripe.api_key = os.getenv("STRIPE_SECRET_KEY", "")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.getenv("FLASK_ENV") == "production"
os.makedirs(app.instance_path, exist_ok=True)
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(app.instance_path, "conges_medflow.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)
BACKUP_DIR = os.path.join(app.instance_path, "backups")
os.makedirs(BACKUP_DIR, exist_ok=True)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024

SCOPES = [
    "openid",
    "email",
    "profile",
    "https://www.googleapis.com/auth/userinfo.email",
    "https://www.googleapis.com/auth/userinfo.profile",
    "https://www.googleapis.com/auth/gmail.send",
    "https://www.googleapis.com/auth/calendar",
    "https://www.googleapis.com/auth/calendar.readonly",
    "https://www.googleapis.com/auth/calendar.events",
]

MONTHS_FR = ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
FIXED_MA_HOLIDAYS = {
    (1,1): "Nouvel An",
    (1,11): "Manifeste de l'Indépendance",
    (1,14): "Nouvel An Amazigh",
    (5,1): "Fête du Travail",
    (7,30): "Fête du Trône",
    (8,14): "Allégeance Oued Eddahab",
    (8,20): "Révolution du Roi et du Peuple",
    (8,21): "Fête de la Jeunesse",
    (11,6): "Marche Verte",
    (11,18): "Fête de l'Indépendance",
}
# Fallback indicatif si Google Calendar n'est pas encore connecté/synchronisé.
HIJRI_FALLBACK = {
    2023: [("2023-04-21","Aïd Al Fitr"),("2023-04-22","Aïd Al Fitr 2"),("2023-06-29","Aïd Al Adha"),("2023-06-30","Aïd Al Adha 2"),("2023-07-19","1er Moharram"),("2023-09-28","Aïd Al Mawlid"),("2023-09-29","Aïd Al Mawlid 2")],
    2024: [("2024-04-10","Aïd Al Fitr"),("2024-04-11","Aïd Al Fitr 2"),("2024-06-17","Aïd Al Adha"),("2024-06-18","Aïd Al Adha 2"),("2024-07-08","1er Moharram"),("2024-09-16","Aïd Al Mawlid"),("2024-09-17","Aïd Al Mawlid 2")],
    2025: [("2025-03-31","Aïd Al Fitr"),("2025-04-01","Aïd Al Fitr 2"),("2025-06-07","Aïd Al Adha"),("2025-06-08","Aïd Al Adha 2"),("2025-06-27","1er Moharram"),("2025-09-05","Aïd Al Mawlid"),("2025-09-06","Aïd Al Mawlid 2")],
    2026: [("2026-03-20","Aïd Al Fitr prévisionnel"),("2026-03-21","Aïd Al Fitr 2 prévisionnel"),("2026-05-27","Aïd Al Adha prévisionnel"),("2026-05-28","Aïd Al Adha 2 prévisionnel"),("2026-06-17","1er Moharram prévisionnel"),("2026-08-25","Aïd Al Mawlid prévisionnel"),("2026-08-26","Aïd Al Mawlid 2 prévisionnel")]
}

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(150), unique=True, nullable=False)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255))
    full_name = db.Column(db.String(150), default="Mohamed AIT ELMALEM")
    company = db.Column(db.String(100), default="MEDFLOW")
    job_title = db.Column(db.String(120), default="Consultant technique SAP Senior")
    hire_date = db.Column(db.Date, default=date(2023,8,1))
    role = db.Column(db.String(20), default="admin")
    google_token = db.Column(db.Text)
    theme = db.Column(db.String(10), default="dark")
    subscription_status = db.Column(db.String(30), default="inactive")
    stripe_customer_id = db.Column(db.String(120))
    stripe_subscription_id = db.Column(db.String(120))

class Holiday(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    day = db.Column(db.Date, unique=True, nullable=False)
    name = db.Column(db.String(150), nullable=False)
    source = db.Column(db.String(30), default="system")
    hijri = db.Column(db.Boolean, default=False)

class LeaveRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    leave_type = db.Column(db.String(50), default="annual")
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    working_days = db.Column(db.Float, default=0)
    status = db.Column(db.String(20), default="pending") # pending/approved/refused
    recipient = db.Column(db.String(150))
    message = db.Column(db.Text)
    decision_comment = db.Column(db.Text)
    medical_note = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    decided_at = db.Column(db.DateTime)
    calendar_event_id = db.Column(db.String(200))

class MonthlyBalance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    credit = db.Column(db.Float, default=1.5)
    extra = db.Column(db.Float, default=0)
    taken = db.Column(db.Float, default=0)
    period = db.Column(db.String(255), default="")
    balance = db.Column(db.Float, default=0)
    locked_seed = db.Column(db.Boolean, default=False)
    __table_args__ = (db.UniqueConstraint("year","month", name="uix_year_month"),)

LEAVE_TYPES = {
    "annual": {"label":"Congé annuel payé", "deduct": True, "default":0},
    "birth": {"label":"Naissance fils/fille", "deduct": False, "default":3},
    "death_parent": {"label":"Décès parent / ascendant", "deduct": False, "default":3},
    "death_spouse_child": {"label":"Décès conjoint / enfant", "deduct": False, "default":3},
    "death_sibling_inlaw": {"label":"Décès frère/sœur/beau-parent", "deduct": False, "default":2},
    "marriage_self": {"label":"Mariage du salarié", "deduct": False, "default":2},
    "marriage_child": {"label":"Mariage d'un enfant", "deduct": False, "default":2},
    "circumcision": {"label":"Circoncision", "deduct": False, "default":2},
    "surgery_family": {"label":"Opération conjoint/enfant à charge", "deduct": False, "default":2},
    "sick": {"label":"Repos maladie", "deduct": False, "default":0},
}

SEED_BALANCES = [
(2023,8,1.5,0,0,"",1.5),(2023,9,1.5,0,0,"",3),(2023,10,1.5,0,0,"",4.5),(2023,11,1.5,0,0,"",6),(2023,12,1.5,0,0,"",7.5),
(2024,1,1.5,0,3,"08-12/01/2024",6),(2024,2,1.5,0,0,"",7.5),(2024,3,1.5,0,0,"",9),(2024,4,1.5,0,4,"05-12/04/2024",6.5),(2024,5,1.5,0,0,"",8),(2024,6,1.5,0,3,"17-21/06/2024",6.5),(2024,7,1.5,0,0,"",8),(2024,8,1.5,0,3,"19-25/08/2024",6.5),(2024,9,1.5,0,5,"23-27/09/2024",3),(2024,10,1.5,0,0,"",4.5),(2024,11,1.5,1,1,"19/11/2024",6),(2024,12,1.5,1,0,"",8.5),
(2025,1,1.5,1,4,"28-31/01/2025",7),(2025,2,1.5,0,0,"",8.5),(2025,3,1.5,0,0,"",10),(2025,4,1.5,0,5,"27/03/2025-07/04/2025",6.5),(2025,5,1.5,0,0,"",8),(2025,6,1.5,0,7,"05-13/06/2025",2.5),(2025,7,1.5,0,3,"09-11/07/2025",1),(2025,8,1.5,0,5,"15-29/08/2025",-2.5),(2025,9,1.5,0,0,"",-1),(2025,10,1.5,0,0,"",0.5),(2025,11,1.5,0,0,"",2),(2025,12,1.5,0,0,"",3.5),
(2026,1,1.5,0,1,"02/01/2026",4),(2026,2,1.5,0,0,"",5.5),(2026,3,1.5,0,2,"09,23/03/2026",5),(2026,4,1.5,0,0,"",6.5),(2026,5,1.5,0,4,"25/05/2026-01/06/2026",4),(2026,6,1.5,0,0,"",5.5),(2026,7,1.5,0,0,"",7),(2026,8,1.5,0,0,"",8.5),(2026,9,1.5,0,0,"",10),(2026,10,1.5,0,0,"",11.5),(2026,11,1.5,0,0,"",13),(2026,12,1.5,0,0,"",14.5),
(2027,1,1.5,0,0,"",16),(2027,2,1.5,0,0,"",17.5),(2027,3,1.5,0,0,"",19),(2027,4,1.5,0,0,"",20.5),(2027,5,1.5,0,0,"",22),(2027,6,1.5,0,0,"",23.5),(2027,7,1.5,0,0,"",25),(2027,8,1.5,0,0,"",26.5),(2027,9,1.5,0,0,"",28),(2027,10,1.5,0,0,"",29.5),(2027,11,1.5,0,0,"",31),(2027,12,1.5,0,0,"",32.5)
]


SEED_APPROVED_REQUESTS = [
    ("annual","2024-01-08","2024-01-12",3,"Congé historique approuvé - 08-12/01/2024"),
    ("annual","2024-04-05","2024-04-12",4,"Congé historique approuvé - 05-12/04/2024"),
    ("annual","2024-06-17","2024-06-21",3,"Congé historique approuvé - 17-21/06/2024"),
    ("annual","2024-08-19","2024-08-25",3,"Congé historique approuvé - 19-25/08/2024"),
    ("annual","2024-09-23","2024-09-27",5,"Congé historique approuvé - 23-27/09/2024"),
    ("annual","2024-11-19","2024-11-19",1,"Congé historique approuvé - 19/11/2024"),
    ("annual","2025-01-28","2025-01-31",4,"Congé historique approuvé - 28-31/01/2025"),
    ("annual","2025-03-27","2025-04-07",5,"Congé historique approuvé - 27/03/2025-07/04/2025"),
    ("annual","2025-06-05","2025-06-13",7,"Congé historique approuvé - 05-13/06/2025"),
    ("annual","2025-07-09","2025-07-11",3,"Congé historique approuvé - 09-11/07/2025"),
    ("annual","2025-08-15","2025-08-29",5,"Congé historique approuvé - 15-29/08/2025"),
    ("annual","2026-01-02","2026-01-02",1,"Congé historique approuvé - 02/01/2026"),
    ("annual","2026-03-09","2026-03-09",1,"Congé historique approuvé - 09/03/2026"),
    ("annual","2026-03-23","2026-03-23",1,"Congé historique approuvé - 23/03/2026"),
    ("annual","2026-05-25","2026-06-01",4,"Congé historique approuvé - 25/05/2026-01/06/2026"),
]


LANGUAGES = {
    "fr": "Français",
    "en": "English",
    "de": "Deutsch",
    "es": "Español",
    "ar": "العربية",
}

TRANSLATIONS = {
    "fr": {
        "dashboard": "Dashboard", "leave_request": "Demande congé", "history": "Historique",
        "calendar": "Calendrier", "holidays": "Jours fériés", "profile": "Profil",
        "subscription": "Abonnement", "guide": "Guide utilisateur", "admin_db": "Admin DB",
        "backups": "Backups", "logout": "Déconnexion", "login": "Connexion",
        "register": "Créer un compte", "language": "Langue", "mode": "Mode",
        "public_home": "Accueil public", "pricing_title": "Abonnement MEDFLOW",
        "secure_payment": "Paiement sécurisé", "trial_7": "7 jours d’essai",
        "welcome": "Bienvenue", "create_request": "Nouvelle demande de congé",
        "year": "Année", "status": "Statut", "actions": "Actions",
        "save": "Enregistrer", "cancel": "Annuler", "approve": "Approuver",
        "refuse": "Refuser", "export_excel": "Export Excel", "export_pdf": "Export PDF",
    },
    "en": {
        "dashboard": "Dashboard", "leave_request": "Leave request", "history": "History",
        "calendar": "Calendar", "holidays": "Holidays", "profile": "Profile",
        "subscription": "Subscription", "guide": "User guide", "admin_db": "Admin DB",
        "backups": "Backups", "logout": "Logout", "login": "Login",
        "register": "Create account", "language": "Language", "mode": "Mode",
        "public_home": "Public home", "pricing_title": "MEDFLOW Subscription",
        "secure_payment": "Secure payment", "trial_7": "7-day trial",
        "welcome": "Welcome", "create_request": "New leave request",
        "year": "Year", "status": "Status", "actions": "Actions",
        "save": "Save", "cancel": "Cancel", "approve": "Approve",
        "refuse": "Reject", "export_excel": "Excel export", "export_pdf": "PDF export",
    },
    "de": {
        "dashboard": "Dashboard", "leave_request": "Urlaubsantrag", "history": "Verlauf",
        "calendar": "Kalender", "holidays": "Feiertage", "profile": "Profil",
        "subscription": "Abonnement", "guide": "Benutzerhandbuch", "admin_db": "Admin DB",
        "backups": "Backups", "logout": "Abmelden", "login": "Anmelden",
        "register": "Konto erstellen", "language": "Sprache", "mode": "Modus",
        "public_home": "Startseite", "pricing_title": "MEDFLOW Abonnement",
        "secure_payment": "Sichere Zahlung", "trial_7": "7 Tage Testphase",
        "welcome": "Willkommen", "create_request": "Neuer Urlaubsantrag",
        "year": "Jahr", "status": "Status", "actions": "Aktionen",
        "save": "Speichern", "cancel": "Abbrechen", "approve": "Genehmigen",
        "refuse": "Ablehnen", "export_excel": "Excel Export", "export_pdf": "PDF Export",
    },
    "es": {
        "dashboard": "Panel", "leave_request": "Solicitud de permiso", "history": "Historial",
        "calendar": "Calendario", "holidays": "Festivos", "profile": "Perfil",
        "subscription": "Suscripción", "guide": "Guía de usuario", "admin_db": "Admin DB",
        "backups": "Copias", "logout": "Cerrar sesión", "login": "Iniciar sesión",
        "register": "Crear cuenta", "language": "Idioma", "mode": "Modo",
        "public_home": "Inicio público", "pricing_title": "Suscripción MEDFLOW",
        "secure_payment": "Pago seguro", "trial_7": "Prueba de 7 días",
        "welcome": "Bienvenido", "create_request": "Nueva solicitud de permiso",
        "year": "Año", "status": "Estado", "actions": "Acciones",
        "save": "Guardar", "cancel": "Cancelar", "approve": "Aprobar",
        "refuse": "Rechazar", "export_excel": "Exportar Excel", "export_pdf": "Exportar PDF",
    },
    "ar": {
        "dashboard": "لوحة التحكم", "leave_request": "طلب عطلة", "history": "السجل",
        "calendar": "التقويم", "holidays": "العطل الرسمية", "profile": "الملف الشخصي",
        "subscription": "الاشتراك", "guide": "دليل المستخدم", "admin_db": "إدارة قاعدة البيانات",
        "backups": "النسخ الاحتياطي", "logout": "تسجيل الخروج", "login": "تسجيل الدخول",
        "register": "إنشاء حساب", "language": "اللغة", "mode": "الوضع",
        "public_home": "الصفحة الرئيسية", "pricing_title": "اشتراك MEDFLOW",
        "secure_payment": "دفع آمن", "trial_7": "تجربة 7 أيام",
        "welcome": "مرحبا", "create_request": "طلب عطلة جديد",
        "year": "السنة", "status": "الحالة", "actions": "الإجراءات",
        "save": "حفظ", "cancel": "إلغاء", "approve": "موافقة",
        "refuse": "رفض", "export_excel": "تصدير Excel", "export_pdf": "تصدير PDF",
    }
}

def get_lang():
    lang = session.get("lang", "fr")
    return lang if lang in LANGUAGES else "fr"

def t(key):
    lang = get_lang()
    return TRANSLATIONS.get(lang, TRANSLATIONS["fr"]).get(key, TRANSLATIONS["fr"].get(key, key))

@app.route("/set_language/<lang>")
def set_language(lang):
    if lang in LANGUAGES:
        session["lang"] = lang
    return redirect(request.referrer or url_for("dashboard"))




FULL_I18N = {
    "fr": {
        "app_name":"MEDFLOW","workspace":"Workspace","owner":"Owner","admin":"Admin","active":"Actif","inactive":"Inactif","trial":"Essai",
        "home":"Accueil","dashboard":"Dashboard","guide":"Guide utilisateur","subscription":"Abonnement","leave_request":"Demande congé","history":"Historique",
        "requests_list":"Liste des demandes","calendar":"Calendrier","holidays":"Jours fériés","profile":"Profil","admin_db":"Admin DB","backups":"Backups",
        "export_excel":"Export Excel","export_pdf":"Export PDF","logout":"Déconnexion","mode":"Mode","language":"Langue","year":"Année","view":"Vue",
        "month":"Mois","credit":"Crédit","extra":"Extra","taken":"Pris","period":"Période","balance":"Solde","status":"Statut","actions":"Actions",
        "created_at":"Créée le","type":"Type","days":"Jours","non_deducted":"Non déduits","deducted":"Déduits","google_calendar":"Google Calendar",
        "approved":"Approuvé","pending":"En attente","refused":"Refusé","cancelled":"Annulé","synchronized":"Synchronisé","not_sync":"Non",
        "new_leave_request":"Nouvelle demande de congé","start_date":"Date début","end_date":"Date fin","recipient_email":"Email destinataire",
        "message":"Message","send_email":"Envoyer email via Gmail API OAuth","save":"Enregistrer","approve":"Approuver","refuse":"Refuser","cancel":"Annuler",
        "annual":"Congé annuel","sick":"Repos maladie","birth":"Naissance","marriage_self":"Mariage salarié","marriage_child":"Mariage enfant",
        "death_family":"Décès famille","operation_family":"Opération conjoint/enfant","medical_certificate":"Certificat médical sécurisé",
        "calendar_standard":"Standard","calendar_hijri":"Hijri","all":"Tous","sync_calendar":"Synchroniser avec calendrier","sync_google":"Synchronisation Google",
        "pricing_title":"Abonnement MEDFLOW","secure_payment":"Paiement sécurisé","trial_7":"7 jours d’essai","monthly":"Mensuel","annual_plan":"Annuel",
        "subscribe_monthly":"S’abonner mensuellement","subscribe_annually":"S’abonner annuellement","use_trial":"Utiliser l’essai","manage_billing":"Gérer facturation",
        "about":"À propos","copyright":"Droits d’auteur","privacy":"Confidentialité","terms":"Conditions","no_data":"Aucune donnée",
        "january":"Janvier","february":"Février","march":"Mars","april":"Avril","may":"Mai","june":"Juin","july":"Juillet","august":"Août","september":"Septembre","october":"Octobre","november":"Novembre","december":"Décembre",
        "monday":"L","tuesday":"M","wednesday":"M","thursday":"J","friday":"V","saturday":"S","sunday":"D"
    },
    "en": {
        "app_name":"MEDFLOW","workspace":"Workspace","owner":"Owner","admin":"Admin","active":"Active","inactive":"Inactive","trial":"Trial",
        "home":"Home","dashboard":"Dashboard","guide":"User guide","subscription":"Subscription","leave_request":"Leave request","history":"History",
        "requests_list":"Requests list","calendar":"Calendar","holidays":"Holidays","profile":"Profile","admin_db":"Admin DB","backups":"Backups",
        "export_excel":"Excel export","export_pdf":"PDF export","logout":"Logout","mode":"Mode","language":"Language","year":"Year","view":"View",
        "month":"Month","credit":"Credit","extra":"Extra","taken":"Taken","period":"Period","balance":"Balance","status":"Status","actions":"Actions",
        "created_at":"Created at","type":"Type","days":"Days","non_deducted":"Non-deducted","deducted":"Deducted","google_calendar":"Google Calendar",
        "approved":"Approved","pending":"Pending","refused":"Rejected","cancelled":"Cancelled","synchronized":"Synced","not_sync":"No",
        "new_leave_request":"New leave request","start_date":"Start date","end_date":"End date","recipient_email":"Recipient email",
        "message":"Message","send_email":"Send email via Gmail API OAuth","save":"Save","approve":"Approve","refuse":"Reject","cancel":"Cancel",
        "annual":"Annual leave","sick":"Sick leave","birth":"Birth","marriage_self":"Employee marriage","marriage_child":"Child marriage",
        "death_family":"Family death","operation_family":"Spouse/child operation","medical_certificate":"Secure medical certificate",
        "calendar_standard":"Standard","calendar_hijri":"Hijri","all":"All","sync_calendar":"Sync with calendar","sync_google":"Google sync",
        "pricing_title":"MEDFLOW subscription","secure_payment":"Secure payment","trial_7":"7-day trial","monthly":"Monthly","annual_plan":"Annual",
        "subscribe_monthly":"Subscribe monthly","subscribe_annually":"Subscribe annually","use_trial":"Use trial","manage_billing":"Manage billing",
        "about":"About","copyright":"Copyright","privacy":"Privacy","terms":"Terms","no_data":"No data",
        "january":"January","february":"February","march":"March","april":"April","may":"May","june":"June","july":"July","august":"August","september":"September","october":"October","november":"November","december":"December",
        "monday":"M","tuesday":"T","wednesday":"W","thursday":"T","friday":"F","saturday":"S","sunday":"S"
    },
    "de": {
        "app_name":"MEDFLOW","workspace":"Arbeitsbereich","owner":"Owner","admin":"Admin","active":"Aktiv","inactive":"Inaktiv","trial":"Test",
        "home":"Startseite","dashboard":"Dashboard","guide":"Benutzerhandbuch","subscription":"Abonnement","leave_request":"Urlaubsantrag","history":"Verlauf",
        "requests_list":"Antragsliste","calendar":"Kalender","holidays":"Feiertage","profile":"Profil","admin_db":"Admin DB","backups":"Backups",
        "export_excel":"Excel Export","export_pdf":"PDF Export","logout":"Abmelden","mode":"Modus","language":"Sprache","year":"Jahr","view":"Ansicht",
        "month":"Monat","credit":"Guthaben","extra":"Extra","taken":"Genommen","period":"Zeitraum","balance":"Saldo","status":"Status","actions":"Aktionen",
        "created_at":"Erstellt am","type":"Typ","days":"Tage","non_deducted":"Nicht abgezogen","deducted":"Abgezogen","google_calendar":"Google Kalender",
        "approved":"Genehmigt","pending":"Ausstehend","refused":"Abgelehnt","cancelled":"Storniert","synchronized":"Synchronisiert","not_sync":"Nein",
        "new_leave_request":"Neuer Urlaubsantrag","start_date":"Startdatum","end_date":"Enddatum","recipient_email":"Empfänger-E-Mail",
        "message":"Nachricht","send_email":"E-Mail über Gmail API OAuth senden","save":"Speichern","approve":"Genehmigen","refuse":"Ablehnen","cancel":"Stornieren",
        "annual":"Jahresurlaub","sick":"Krankschreibung","birth":"Geburt","marriage_self":"Heirat Mitarbeiter","marriage_child":"Heirat Kind",
        "death_family":"Todesfall Familie","operation_family":"Operation Ehepartner/Kind","medical_certificate":"Sicheres ärztliches Attest",
        "calendar_standard":"Standard","calendar_hijri":"Hijri","all":"Alle","sync_calendar":"Mit Kalender synchronisieren","sync_google":"Google-Synchronisierung",
        "pricing_title":"MEDFLOW Abonnement","secure_payment":"Sichere Zahlung","trial_7":"7 Tage Testphase","monthly":"Monatlich","annual_plan":"Jährlich",
        "subscribe_monthly":"Monatlich abonnieren","subscribe_annually":"Jährlich abonnieren","use_trial":"Testphase nutzen","manage_billing":"Abrechnung verwalten",
        "about":"Über uns","copyright":"Urheberrecht","privacy":"Datenschutz","terms":"Bedingungen","no_data":"Keine Daten",
        "january":"Januar","february":"Februar","march":"März","april":"April","may":"Mai","june":"Juni","july":"Juli","august":"August","september":"September","october":"Oktober","november":"November","december":"Dezember",
        "monday":"Mo","tuesday":"Di","wednesday":"Mi","thursday":"Do","friday":"Fr","saturday":"Sa","sunday":"So"
    },
    "es": {
        "app_name":"MEDFLOW","workspace":"Espacio","owner":"Owner","admin":"Admin","active":"Activo","inactive":"Inactivo","trial":"Prueba",
        "home":"Inicio","dashboard":"Panel","guide":"Guía de usuario","subscription":"Suscripción","leave_request":"Solicitud de permiso","history":"Historial",
        "requests_list":"Lista de solicitudes","calendar":"Calendario","holidays":"Festivos","profile":"Perfil","admin_db":"Admin DB","backups":"Copias",
        "export_excel":"Exportar Excel","export_pdf":"Exportar PDF","logout":"Cerrar sesión","mode":"Modo","language":"Idioma","year":"Año","view":"Vista",
        "month":"Mes","credit":"Crédito","extra":"Extra","taken":"Tomado","period":"Periodo","balance":"Saldo","status":"Estado","actions":"Acciones",
        "created_at":"Creado el","type":"Tipo","days":"Días","non_deducted":"No deducidos","deducted":"Deducidos","google_calendar":"Google Calendar",
        "approved":"Aprobado","pending":"Pendiente","refused":"Rechazado","cancelled":"Cancelado","synchronized":"Sincronizado","not_sync":"No",
        "new_leave_request":"Nueva solicitud de permiso","start_date":"Fecha inicio","end_date":"Fecha fin","recipient_email":"Email destinatario",
        "message":"Mensaje","send_email":"Enviar email vía Gmail API OAuth","save":"Guardar","approve":"Aprobar","refuse":"Rechazar","cancel":"Cancelar",
        "annual":"Permiso anual","sick":"Baja médica","birth":"Nacimiento","marriage_self":"Matrimonio empleado","marriage_child":"Matrimonio hijo",
        "death_family":"Fallecimiento familiar","operation_family":"Operación cónyuge/hijo","medical_certificate":"Certificado médico seguro",
        "calendar_standard":"Estándar","calendar_hijri":"Hijri","all":"Todos","sync_calendar":"Sincronizar con calendario","sync_google":"Sincronización Google",
        "pricing_title":"Suscripción MEDFLOW","secure_payment":"Pago seguro","trial_7":"Prueba de 7 días","monthly":"Mensual","annual_plan":"Anual",
        "subscribe_monthly":"Suscribirse mensual","subscribe_annually":"Suscribirse anual","use_trial":"Usar prueba","manage_billing":"Gestionar facturación",
        "about":"Sobre nosotros","copyright":"Derechos de autor","privacy":"Privacidad","terms":"Términos","no_data":"Sin datos",
        "january":"Enero","february":"Febrero","march":"Marzo","april":"Abril","may":"Mayo","june":"Junio","july":"Julio","august":"Agosto","september":"Septiembre","october":"Octubre","november":"Noviembre","december":"Diciembre",
        "monday":"L","tuesday":"M","wednesday":"X","thursday":"J","friday":"V","saturday":"S","sunday":"D"
    },
    "ar": {
        "app_name":"MEDFLOW","workspace":"مساحة العمل","owner":"المالك","admin":"المدير","active":"نشط","inactive":"غير نشط","trial":"تجربة",
        "home":"الرئيسية","dashboard":"لوحة التحكم","guide":"دليل المستخدم","subscription":"الاشتراك","leave_request":"طلب عطلة","history":"السجل",
        "requests_list":"قائمة الطلبات","calendar":"التقويم","holidays":"العطل الرسمية","profile":"الملف الشخصي","admin_db":"إدارة قاعدة البيانات","backups":"النسخ الاحتياطي",
        "export_excel":"تصدير Excel","export_pdf":"تصدير PDF","logout":"تسجيل الخروج","mode":"الوضع","language":"اللغة","year":"السنة","view":"العرض",
        "month":"الشهر","credit":"الرصيد","extra":"إضافي","taken":"المأخوذ","period":"الفترة","balance":"الرصيد المتبقي","status":"الحالة","actions":"الإجراءات",
        "created_at":"تاريخ الإنشاء","type":"النوع","days":"الأيام","non_deducted":"غير مخصومة","deducted":"مخصومة","google_calendar":"تقويم Google",
        "approved":"موافق عليه","pending":"قيد الانتظار","refused":"مرفوض","cancelled":"ملغى","synchronized":"متزامن","not_sync":"لا",
        "new_leave_request":"طلب عطلة جديد","start_date":"تاريخ البداية","end_date":"تاريخ النهاية","recipient_email":"بريد المستلم",
        "message":"الرسالة","send_email":"إرسال عبر Gmail API OAuth","save":"حفظ","approve":"موافقة","refuse":"رفض","cancel":"إلغاء",
        "annual":"عطلة سنوية","sick":"راحة مرضية","birth":"ازدياد","marriage_self":"زواج الموظف","marriage_child":"زواج الابن/الابنة",
        "death_family":"وفاة في العائلة","operation_family":"عملية للزوج/الطفل","medical_certificate":"شهادة طبية آمنة",
        "calendar_standard":"عادي","calendar_hijri":"هجري","all":"الكل","sync_calendar":"مزامنة مع التقويم","sync_google":"مزامنة Google",
        "pricing_title":"اشتراك MEDFLOW","secure_payment":"دفع آمن","trial_7":"تجربة 7 أيام","monthly":"شهري","annual_plan":"سنوي",
        "subscribe_monthly":"اشترك شهرياً","subscribe_annually":"اشترك سنوياً","use_trial":"استخدم التجربة","manage_billing":"إدارة الفوترة",
        "about":"من نحن","copyright":"حقوق النشر","privacy":"الخصوصية","terms":"الشروط","no_data":"لا توجد بيانات",
        "january":"يناير","february":"فبراير","march":"مارس","april":"أبريل","may":"ماي","june":"يونيو","july":"يوليوز","august":"غشت","september":"شتنبر","october":"أكتوبر","november":"نونبر","december":"دجنبر",
        "monday":"ن","tuesday":"ث","wednesday":"ر","thursday":"خ","friday":"ج","saturday":"س","sunday":"ح"
    }
}

def tt(key):
    try:
        lang = get_lang()
    except Exception:
        lang = "fr"
    return FULL_I18N.get(lang, FULL_I18N["fr"]).get(key, key)

def month_name_i18n(m):
    keys = ["january","february","march","april","may","june","july","august","september","october","november","december"]
    return tt(keys[int(m)-1])

def status_i18n(s):
    return tt(str(s or "").lower())

def patch_full_i18n():
    extra = {
        "fr": {
            "login_title":"Connexion sécurisée","register_title":"Créer un compte","full_name":"Nom complet","email":"Email","username":"Nom utilisateur","password":"Mot de passe","confirm_password":"Confirmer mot de passe","company":"Entreprise","job_title":"Poste","hire_date":"Date d'embauche",
            "admin_only":"Admin uniquement","restore_backup":"Restaurer backup","upload_backup":"Uploader un backup","create_backup":"Créer un backup","download":"Télécharger","restore":"Restaurer","delete":"Supprimer","edit":"Modifier","security":"Sécurité","settings":"Paramètres",
            "leave_balance":"Solde congé","approved_leaves":"Congés approuvés","remaining":"Restant","used":"Utilisé","google_connected":"Google connecté","google_not_connected":"Google non connecté","connect_google":"Connecter Google","disconnect_google":"Déconnecter Google",
            "help_title":"Guide utilisateur MEDFLOW","privacy_title":"Politique de confidentialité","terms_title":"Conditions d’utilisation","about_title":"À propos de nous","copyright_title":"Droits d’auteur",
            "no_card_stored":"Aucune carte bancaire n’est stockée sur MEDFLOW.","stripe_secure":"Paiement sécurisé via Stripe Checkout.","all_rights":"Tous droits réservés.","back":"Retour","search":"Rechercher","filter":"Filtrer"
        },
        "en": {
            "login_title":"Secure login","register_title":"Create account","full_name":"Full name","email":"Email","username":"Username","password":"Password","confirm_password":"Confirm password","company":"Company","job_title":"Job title","hire_date":"Hire date",
            "admin_only":"Admin only","restore_backup":"Restore backup","upload_backup":"Upload backup","create_backup":"Create backup","download":"Download","restore":"Restore","delete":"Delete","edit":"Edit","security":"Security","settings":"Settings",
            "leave_balance":"Leave balance","approved_leaves":"Approved leaves","remaining":"Remaining","used":"Used","google_connected":"Google connected","google_not_connected":"Google not connected","connect_google":"Connect Google","disconnect_google":"Disconnect Google",
            "help_title":"MEDFLOW user guide","privacy_title":"Privacy policy","terms_title":"Terms of use","about_title":"About us","copyright_title":"Copyright",
            "no_card_stored":"No card data is stored on MEDFLOW.","stripe_secure":"Secure payment via Stripe Checkout.","all_rights":"All rights reserved.","back":"Back","search":"Search","filter":"Filter"
        },
        "de": {
            "login_title":"Sichere Anmeldung","register_title":"Konto erstellen","full_name":"Vollständiger Name","email":"E-Mail","username":"Benutzername","password":"Passwort","confirm_password":"Passwort bestätigen","company":"Unternehmen","job_title":"Position","hire_date":"Einstellungsdatum",
            "admin_only":"Nur Admin","restore_backup":"Backup wiederherstellen","upload_backup":"Backup hochladen","create_backup":"Backup erstellen","download":"Herunterladen","restore":"Wiederherstellen","delete":"Löschen","edit":"Bearbeiten","security":"Sicherheit","settings":"Einstellungen",
            "leave_balance":"Urlaubssaldo","approved_leaves":"Genehmigte Urlaube","remaining":"Verbleibend","used":"Verwendet","google_connected":"Google verbunden","google_not_connected":"Google nicht verbunden","connect_google":"Google verbinden","disconnect_google":"Google trennen",
            "help_title":"MEDFLOW Benutzerhandbuch","privacy_title":"Datenschutzerklärung","terms_title":"Nutzungsbedingungen","about_title":"Über uns","copyright_title":"Urheberrecht",
            "no_card_stored":"Kartendaten werden nicht auf MEDFLOW gespeichert.","stripe_secure":"Sichere Zahlung über Stripe Checkout.","all_rights":"Alle Rechte vorbehalten.","back":"Zurück","search":"Suchen","filter":"Filtern"
        },
        "es": {
            "login_title":"Inicio de sesión seguro","register_title":"Crear cuenta","full_name":"Nombre completo","email":"Email","username":"Usuario","password":"Contraseña","confirm_password":"Confirmar contraseña","company":"Empresa","job_title":"Puesto","hire_date":"Fecha de contratación",
            "admin_only":"Solo admin","restore_backup":"Restaurar backup","upload_backup":"Subir backup","create_backup":"Crear backup","download":"Descargar","restore":"Restaurar","delete":"Eliminar","edit":"Editar","security":"Seguridad","settings":"Ajustes",
            "leave_balance":"Saldo de permisos","approved_leaves":"Permisos aprobados","remaining":"Restante","used":"Usado","google_connected":"Google conectado","google_not_connected":"Google no conectado","connect_google":"Conectar Google","disconnect_google":"Desconectar Google",
            "help_title":"Guía de usuario MEDFLOW","privacy_title":"Política de privacidad","terms_title":"Condiciones de uso","about_title":"Sobre nosotros","copyright_title":"Derechos de autor",
            "no_card_stored":"No se almacenan datos de tarjeta en MEDFLOW.","stripe_secure":"Pago seguro vía Stripe Checkout.","all_rights":"Todos los derechos reservados.","back":"Volver","search":"Buscar","filter":"Filtrar"
        },
        "ar": {
            "login_title":"تسجيل دخول آمن","register_title":"إنشاء حساب","full_name":"الاسم الكامل","email":"البريد الإلكتروني","username":"اسم المستخدم","password":"كلمة المرور","confirm_password":"تأكيد كلمة المرور","company":"الشركة","job_title":"المنصب","hire_date":"تاريخ التوظيف",
            "admin_only":"للمدير فقط","restore_backup":"استعادة نسخة احتياطية","upload_backup":"رفع نسخة احتياطية","create_backup":"إنشاء نسخة احتياطية","download":"تحميل","restore":"استعادة","delete":"حذف","edit":"تعديل","security":"الأمان","settings":"الإعدادات",
            "leave_balance":"رصيد العطل","approved_leaves":"العطل الموافق عليها","remaining":"المتبقي","used":"المستخدم","google_connected":"Google متصل","google_not_connected":"Google غير متصل","connect_google":"ربط Google","disconnect_google":"فصل Google",
            "help_title":"دليل مستخدم MEDFLOW","privacy_title":"سياسة الخصوصية","terms_title":"شروط الاستخدام","about_title":"من نحن","copyright_title":"حقوق النشر",
            "no_card_stored":"لا يتم تخزين بيانات البطاقة في MEDFLOW.","stripe_secure":"دفع آمن عبر Stripe Checkout.","all_rights":"جميع الحقوق محفوظة.","back":"رجوع","search":"بحث","filter":"تصفية"
        }
    }
    try:
        for lang, values in extra.items():
            FULL_I18N.setdefault(lang, {})
            FULL_I18N[lang].update(values)
    except Exception:
        pass

patch_full_i18n()


def safe_t(key):
    try:
        return t(key)
    except Exception:
        fallback = {
            "public_home": "Accueil public",
            "dashboard": "Dashboard",
            "guide": "Guide utilisateur",
            "subscription": "Abonnement",
            "leave_request": "Demande congé",
            "history": "Historique",
            "calendar": "Calendrier",
            "holidays": "Jours fériés",
            "profile": "Profil",
            "admin_db": "Admin DB",
            "backups": "Backups",
            "export_excel": "Export Excel",
            "export_pdf": "Export PDF",
            "logout": "Déconnexion",
            "language": "Langue",
            "mode": "Mode",
        }
        return fallback.get(key, key)



LEAVE_TYPE_I18N = {
    "annual": {"fr":"Congé annuel payé","en":"Paid annual leave","de":"Bezahlter Jahresurlaub","es":"Vacaciones anuales pagadas","ar":"عطلة سنوية مدفوعة"},
    "birth": {"fr":"Naissance fils/fille","en":"Birth of child","de":"Geburt eines Kindes","es":"Nacimiento de hijo/a","ar":"ازدياد ابن/ابنة"},
    "death_parent": {"fr":"Décès parent / ascendant","en":"Death of parent / ascendant","de":"Tod Elternteil / Vorfahr","es":"Fallecimiento padre/madre/ascendiente","ar":"وفاة أحد الوالدين / الأصول"},
    "death_spouse_child": {"fr":"Décès conjoint / enfant","en":"Death of spouse / child","de":"Tod Ehepartner / Kind","es":"Fallecimiento cónyuge / hijo","ar":"وفاة الزوج/الزوجة أو الابن"},
    "death_sibling_parent_inlaw": {"fr":"Décès frère/sœur/beau-parent","en":"Death of sibling / parent-in-law","de":"Tod Geschwister / Schwiegereltern","es":"Fallecimiento hermano/a / suegro/a","ar":"وفاة أخ/أخت أو أحد الأصهار"},
    "marriage_self": {"fr":"Mariage du salarié","en":"Employee marriage","de":"Heirat des Mitarbeiters","es":"Matrimonio del empleado","ar":"زواج الموظف"},
    "marriage_child": {"fr":"Mariage d'un enfant","en":"Child marriage","de":"Heirat eines Kindes","es":"Matrimonio de un hijo","ar":"زواج ابن/ابنة"},
    "circumcision": {"fr":"Circoncision","en":"Circumcision","de":"Beschneidung","es":"Circuncisión","ar":"الختان"},
    "operation_family": {"fr":"Opération conjoint/enfant à charge","en":"Spouse/dependent child operation","de":"Operation Ehepartner/unterhaltsberechtigtes Kind","es":"Operación cónyuge/hijo a cargo","ar":"عملية للزوج/الطفل المكفول"},
    "sick": {"fr":"Repos maladie","en":"Sick leave","de":"Krankschreibung","es":"Baja médica","ar":"راحة مرضية"},
}

def leave_label_i18n(key):
    try:
        lang = get_lang()
    except Exception:
        lang = "fr"
    return LEAVE_TYPE_I18N.get(key, {}).get(lang, LEAVE_TYPES.get(key, {}).get("label", key))


def patch_v31_i18n():
    extra = {
        "fr": {
            "leave_management":"Gestion des congés","privacy_first":"Confidentialité d’abord","multi_language":"Multilingue",
            "about_desc":"MEDFLOW est une plateforme moderne pour gérer les congés, absences, calendriers et abonnements de manière simple et sécurisée.",
            "our_mission":"Notre mission","mission_desc":"Offrir un outil clair, accessible et sécurisé pour les utilisateurs individuels, freelances et petites équipes.",
            "request_intro":"Créez une demande claire, calculez les jours ouvrables et envoyez-la depuis votre compte Google.",
            "email_message_custom":"Message email personnalisé","latest_requests":"Dernières demandes","direct_tracking":"Suivi direct",
            "deduct_from_balance":"À déduire du solde","estimated_working_days":"Jours ouvrables estimés","authorized_days":"Jours autorisés occasion / loi",
            "recipient_email":"Email destinataire","leave_type":"Type de congé","period_start":"Date début","period_end":"Date fin"
        },
        "en": {
            "leave_management":"Leave Management","privacy_first":"Privacy First","multi_language":"Multi-language",
            "about_desc":"MEDFLOW is a modern platform for managing leave, absences, calendars and subscriptions simply and securely.",
            "our_mission":"Our mission","mission_desc":"Provide a clear, accessible and secure tool for individuals, freelancers and small teams.",
            "request_intro":"Create a clear request, calculate working days and send it from your Google account.",
            "email_message_custom":"Custom email message","latest_requests":"Latest requests","direct_tracking":"Direct tracking",
            "deduct_from_balance":"To deduct from balance","estimated_working_days":"Estimated working days","authorized_days":"Authorized occasion / legal days",
            "recipient_email":"Recipient email","leave_type":"Leave type","period_start":"Start date","period_end":"End date"
        },
        "de": {
            "leave_management":"Urlaubsverwaltung","privacy_first":"Datenschutz zuerst","multi_language":"Mehrsprachig",
            "about_desc":"MEDFLOW ist eine moderne Plattform zur einfachen und sicheren Verwaltung von Urlaub, Abwesenheiten, Kalendern und Abonnements.",
            "our_mission":"Unsere Mission","mission_desc":"Ein klares, zugängliches und sicheres Tool für Einzelpersonen, Freelancer und kleine Teams bereitstellen.",
            "request_intro":"Erstellen Sie einen klaren Antrag, berechnen Sie Arbeitstage und senden Sie ihn über Ihr Google-Konto.",
            "email_message_custom":"Individuelle E-Mail-Nachricht","latest_requests":"Letzte Anträge","direct_tracking":"Direkte Verfolgung",
            "deduct_from_balance":"Vom Saldo abzuziehen","estimated_working_days":"Geschätzte Arbeitstage","authorized_days":"Genehmigte Anlass-/Gesetzestage",
            "recipient_email":"Empfänger-E-Mail","leave_type":"Urlaubsart","period_start":"Startdatum","period_end":"Enddatum"
        },
        "es": {
            "leave_management":"Gestión de permisos","privacy_first":"Privacidad primero","multi_language":"Multilingüe",
            "about_desc":"MEDFLOW es una plataforma moderna para gestionar permisos, ausencias, calendarios y suscripciones de forma simple y segura.",
            "our_mission":"Nuestra misión","mission_desc":"Ofrecer una herramienta clara, accesible y segura para usuarios individuales, freelancers y equipos pequeños.",
            "request_intro":"Crea una solicitud clara, calcula los días laborables y envíala desde tu cuenta Google.",
            "email_message_custom":"Mensaje email personalizado","latest_requests":"Últimas solicitudes","direct_tracking":"Seguimiento directo",
            "deduct_from_balance":"A deducir del saldo","estimated_working_days":"Días laborables estimados","authorized_days":"Días autorizados / legales",
            "recipient_email":"Email destinatario","leave_type":"Tipo de permiso","period_start":"Fecha inicio","period_end":"Fecha fin"
        },
        "ar": {
            "leave_management":"تدبير العطل","privacy_first":"الخصوصية أولاً","multi_language":"متعدد اللغات",
            "about_desc":"MEDFLOW منصة حديثة لتدبير العطل والغيابات والتقويمات والاشتراكات بطريقة بسيطة وآمنة.",
            "our_mission":"مهمتنا","mission_desc":"تقديم أداة واضحة وآمنة وسهلة الاستخدام للأفراد والمستقلين والفرق الصغيرة.",
            "request_intro":"أنشئ طلباً واضحاً، احسب أيام العمل وأرسله من حساب Google الخاص بك.",
            "email_message_custom":"رسالة بريد مخصصة","latest_requests":"آخر الطلبات","direct_tracking":"تتبع مباشر",
            "deduct_from_balance":"سيتم خصمها من الرصيد","estimated_working_days":"أيام العمل المقدرة","authorized_days":"الأيام المرخصة / القانونية",
            "recipient_email":"بريد المستلم","leave_type":"نوع العطلة","period_start":"تاريخ البداية","period_end":"تاريخ النهاية"
        }
    }
    try:
        for lang, values in extra.items():
            FULL_I18N.setdefault(lang, {})
            FULL_I18N[lang].update(values)
    except Exception:
        pass

patch_v31_i18n()


# -------------------------
# V33 Full i18n hardcoded text map + profiling
# -------------------------
I18N_TEXT_MAP = {
    "fr": {},
    "en": {
        "Gestion congés secteur privé Maroc": "Morocco private sector leave management",
        "Suivi annuel, ancienneté, jours fériés Maroc, Hijri, approbation/refus et export.": "Annual tracking, seniority, Morocco holidays, Hijri, approval/rejection and export.",
        "Bilan mensuel": "Monthly summary",
        "Mini panneau Admin DB": "Mini Admin DB panel",
        "Voir, modifier ou supprimer les données SQLite depuis Render. Accès réservé admin. Champs sensibles masqués.": "View, edit or delete SQLite data from Render. Admin access only. Sensitive fields hidden.",
        "Administration sécurisée": "Secure administration",
        "Centre d’aide": "Help center",
        "Tout ce qu’il faut savoir pour utiliser l’application simplement et en sécurité.": "Everything you need to use the app simply and securely.",
        "Connexion": "Login",
        "Crée ton compte ou connecte-toi avec Google. Pour envoyer des emails, connecte Google depuis le profil.": "Create your account or sign in with Google. To send emails, connect Google from your profile.",
        "Demande congé": "Leave request",
        "Choisis le type, la période et le destinataire. Les week-ends et jours fériés ne sont pas comptés.": "Choose the type, period and recipient. Weekends and holidays are not counted.",
        "Occasions": "Occasions",
        "Pour naissance, mariage ou décès, renseigne les jours autorisés non déduits. Le reste est calculé automatiquement.": "For birth, marriage or death, enter authorized non-deducted days. The rest is calculated automatically.",
        "Maladie": "Sickness",
        "Pour repos maladie, renseigne une référence/certificat. Évite de saisir des données médicales sensibles inutiles.": "For sick leave, enter a reference/certificate. Avoid entering unnecessary sensitive medical data.",
        "Google Calendar": "Google Calendar",
        "Un congé approuvé peut être synchronisé dans Google Calendar. En cas d’annulation, l’événement est supprimé.": "An approved leave can be synced to Google Calendar. If cancelled, the event is removed.",
        "Exports": "Exports",
        "Tu peux télécharger les bilans en Excel et PDF depuis le menu.": "You can download summaries in Excel and PDF from the menu.",
        "Abonnement": "Subscription",
        "Les utilisateurs non-admin doivent avoir un abonnement actif pour accéder aux fonctions principales.": "Non-admin users need an active subscription to access main features.",
        "Sécurité": "Security",
        "Utilise un mot de passe fort, garde ton compte Google sécurisé et ne partage jamais tes accès.": "Use a strong password, keep your Google account secure and never share your access.",
        "Données séparées": "Separated data",
        "Les données de Mohamed AIT ELMALEM appartiennent au compte owner. Chaque nouvel utilisateur commence à zéro et ajoute ses propres congés.": "The owner's data belongs to the owner account. Each new user starts from zero and adds their own leaves.",
        "Nouvelle demande de congé": "New leave request",
        "Portail collaborateur": "Employee portal",
        "Choisis la période avec calendrier. Les week-ends et jours fériés ne sont pas comptés.": "Choose the period using the calendar. Weekends and holidays are not counted.",
        "Créer la demande": "Create request",
        "Dernières demandes": "Latest requests",
        "Synchroniser avec calendrier": "Sync with calendar",
        "Toutes les demandes": "All requests",
        "Synchronisation depuis Google Calendar": "Sync from Google Calendar",
        "Cette page affiche toutes les demandes et peut importer les congés déjà présents dans Google Calendar.": "This page shows all requests and can import leaves already present in Google Calendar.",
        "Import congés existants": "Import existing leaves",
        "Synchroniser avec calendrier": "Sync with calendar",
        "Depuis": "From",
        "Jusqu’à": "To",
        "Calendriers Maroc : Standard & Hijri": "Morocco calendars: Standard & Hijri",
        "Sélectionne le calendrier à synchroniser depuis Google Calendar : jours fériés standards ou événements Hijri.": "Select the calendar to sync from Google Calendar: standard holidays or Hijri events.",
        "Profil": "Profile",
        "État Google": "Google status",
        "Google non connecté. Connecte-toi avec Google pour envoyer les emails depuis ton adresse.": "Google not connected. Connect with Google to send emails from your address.",
        "Connecter Google": "Connect Google",
        "Enregistrer": "Save",
        "Backups base de données": "Database backups",
        "Protection des données": "Data protection",
        "Crée, télécharge et restaure des sauvegardes SQLite. Accès réservé admin.": "Create, download and restore SQLite backups. Admin only.",
        "Stratégie recommandée": "Recommended strategy",
        "Créer un backup maintenant": "Create backup now",
        "Fichiers backup": "Backup files",
        "Uploader un backup": "Upload backup",
        "Restore sécurisé": "Secure restore",
        "Seuls les fichiers SQLite sont autorisés.": "Only SQLite files are allowed.",
        "Historique des demandes": "Request history",
        "Traçabilité complète": "Full traceability",
        "Les congés déjà pris sont chargés ici comme demandes approuvées, avec période, jours pris, statut et commentaire.": "Previously taken leaves are loaded here as approved requests with period, days, status and comment.",
        "Liste des demandes": "Request list",
        "Demandes & Google Calendar": "Requests & Google Calendar",
        "Annuler": "Cancel",
        "Motif annulation": "Cancellation reason",
        "Synchronisé depuis le bilan fourni": "Synced from provided summary",
        "Congé historique approuvé": "Historical approved leave",
        "Congé annuel payé": "Paid annual leave",
        "Mini panneau Admin DB": "Mini Admin DB panel",
        "Profiling": "Profiling",
        "Performances": "Performance"
    },
    "de": {
        "Gestion congés secteur privé Maroc": "Urlaubsverwaltung Privatsektor Marokko",
        "Suivi annuel, ancienneté, jours fériés Maroc, Hijri, approbation/refus et export.": "Jährliche Nachverfolgung, Dienstalter, Feiertage Marokko, Hijri, Genehmigung/Ablehnung und Export.",
        "Bilan mensuel": "Monatsübersicht",
        "Mini panneau Admin DB": "Mini Admin-DB-Panel",
        "Voir, modifier ou supprimer les données SQLite depuis Render. Accès réservé admin. Champs sensibles masqués.": "SQLite-Daten von Render anzeigen, bearbeiten oder löschen. Nur Admin. Sensible Felder sind verborgen.",
        "Administration sécurisée": "Sichere Administration",
        "Centre d’aide": "Hilfezentrum",
        "Tout ce qu’il faut savoir pour utiliser l’application simplement et en sécurité.": "Alles, was Sie brauchen, um die App einfach und sicher zu nutzen.",
        "Connexion": "Anmeldung",
        "Crée ton compte ou connecte-toi avec Google. Pour envoyer des emails, connecte Google depuis le profil.": "Erstellen Sie ein Konto oder melden Sie sich mit Google an. Zum Senden von E-Mails verbinden Sie Google im Profil.",
        "Demande congé": "Urlaubsantrag",
        "Choisis le type, la période et le destinataire. Les week-ends et jours fériés ne sont pas comptés.": "Wählen Sie Typ, Zeitraum und Empfänger. Wochenenden und Feiertage werden nicht gezählt.",
        "Occasions": "Anlässe",
        "Pour naissance, mariage ou décès, renseigne les jours autorisés non déduits. Le reste est calculé automatiquement.": "Für Geburt, Hochzeit oder Todesfall geben Sie autorisierte nicht abgezogene Tage ein. Der Rest wird automatisch berechnet.",
        "Maladie": "Krankheit",
        "Pour repos maladie, renseigne une référence/certificat. Évite de saisir des données médicales sensibles inutiles.": "Für Krankheitsurlaub geben Sie eine Referenz/ein Attest ein. Vermeiden Sie unnötige sensible medizinische Daten.",
        "Un congé approuvé peut être synchronisé dans Google Calendar. En cas d’annulation, l’événement est supprimé.": "Genehmigter Urlaub kann mit Google Calendar synchronisiert werden. Bei Stornierung wird das Ereignis gelöscht.",
        "Tu peux télécharger les bilans en Excel et PDF depuis le menu.": "Sie können Berichte im Menü als Excel und PDF herunterladen.",
        "Les utilisateurs non-admin doivent avoir un abonnement actif pour accéder aux fonctions principales.": "Nicht-Admin-Benutzer benötigen ein aktives Abonnement für die Hauptfunktionen.",
        "Sécurité": "Sicherheit",
        "Utilise un mot de passe fort, garde ton compte Google sécurisé et ne partage jamais tes accès.": "Verwenden Sie ein starkes Passwort, sichern Sie Ihr Google-Konto und teilen Sie niemals Ihre Zugänge.",
        "Données séparées": "Getrennte Daten",
        "Les données de Mohamed AIT ELMALEM appartiennent au compte owner. Chaque nouvel utilisateur commence à zéro et ajoute ses propres congés.": "Die Owner-Daten gehören zum Owner-Konto. Jeder neue Benutzer beginnt bei null und fügt eigene Urlaube hinzu.",
        "Nouvelle demande de congé": "Neuer Urlaubsantrag",
        "Portail collaborateur": "Mitarbeiterportal",
        "Choisis la période avec calendrier. Les week-ends et jours fériés ne sont pas comptés.": "Wählen Sie den Zeitraum im Kalender. Wochenenden und Feiertage werden nicht gezählt.",
        "Créer la demande": "Antrag erstellen",
        "Dernières demandes": "Letzte Anträge",
        "Synchroniser avec calendrier": "Mit Kalender synchronisieren",
        "Toutes les demandes": "Alle Anträge",
        "Synchronisation depuis Google Calendar": "Synchronisierung aus Google Calendar",
        "Cette page affiche toutes les demandes et peut importer les congés déjà présents dans Google Calendar.": "Diese Seite zeigt alle Anträge und kann bereits vorhandene Urlaube aus Google Calendar importieren.",
        "Import congés existants": "Bestehende Urlaube importieren",
        "Depuis": "Von",
        "Jusqu’à": "Bis",
        "Calendriers Maroc : Standard & Hijri": "Kalender Marokko: Standard & Hijri",
        "Sélectionne le calendrier à synchroniser depuis Google Calendar : jours fériés standards ou événements Hijri.": "Wählen Sie den Kalender zur Synchronisierung aus Google Calendar: Standard-Feiertage oder Hijri-Ereignisse.",
        "Profil": "Profil",
        "État Google": "Google-Status",
        "Google non connecté. Connecte-toi avec Google pour envoyer les emails depuis ton adresse.": "Google nicht verbunden. Verbinden Sie Google, um E-Mails von Ihrer Adresse zu senden.",
        "Connecter Google": "Google verbinden",
        "Enregistrer": "Speichern",
        "Backups base de données": "Datenbank-Backups",
        "Protection des données": "Datenschutz",
        "Crée, télécharge et restaure des sauvegardes SQLite. Accès réservé admin.": "SQLite-Backups erstellen, herunterladen und wiederherstellen. Nur Admin.",
        "Stratégie recommandée": "Empfohlene Strategie",
        "Créer un backup maintenant": "Backup jetzt erstellen",
        "Fichiers backup": "Backup-Dateien",
        "Uploader un backup": "Backup hochladen",
        "Restore sécurisé": "Sichere Wiederherstellung",
        "Seuls les fichiers SQLite sont autorisés.": "Nur SQLite-Dateien sind erlaubt.",
        "Historique des demandes": "Antragshistorie",
        "Traçabilité complète": "Vollständige Nachverfolgbarkeit",
        "Les congés déjà pris sont chargés ici comme demandes approuvées, avec période, jours pris, statut et commentaire.": "Bereits genommene Urlaube werden hier als genehmigte Anträge mit Zeitraum, Tagen, Status und Kommentar geladen.",
        "Liste des demandes": "Antragsliste",
        "Demandes & Google Calendar": "Anträge & Google Calendar",
        "Annuler": "Stornieren",
        "Motif annulation": "Stornierungsgrund",
        "Synchronisé depuis le bilan fourni": "Aus bereitgestellter Übersicht synchronisiert",
        "Congé historique approuvé": "Historisch genehmigter Urlaub",
        "Congé annuel payé": "Bezahlter Jahresurlaub",
        "Profiling": "Profiling",
        "Performances": "Performance"
    },
    "es": {},
    "ar": {
        "Gestion congés secteur privé Maroc": "تدبير العطل للقطاع الخاص بالمغرب",
        "Suivi annuel, ancienneté, jours fériés Maroc, Hijri, approbation/refus et export.": "تتبع سنوي، الأقدمية، العطل الرسمية بالمغرب، الهجري، الموافقة/الرفض والتصدير.",
        "Bilan mensuel": "الحصيلة الشهرية",
        "Mini panneau Admin DB": "لوحة إدارة قاعدة البيانات",
        "Voir, modifier ou supprimer les données SQLite depuis Render. Accès réservé admin. Champs sensibles masqués.": "عرض أو تعديل أو حذف بيانات SQLite من Render. الوصول للمدير فقط. الحقول الحساسة مخفية.",
        "Administration sécurisée": "إدارة آمنة",
        "Centre d’aide": "مركز المساعدة",
        "Tout ce qu’il faut savoir pour utiliser l’application simplement et en sécurité.": "كل ما تحتاجه لاستخدام التطبيق بسهولة وأمان.",
        "Connexion": "تسجيل الدخول",
        "Crée ton compte ou connecte-toi avec Google. Pour envoyer des emails, connecte Google depuis le profil.": "أنشئ حسابك أو سجل الدخول عبر Google. لإرسال الرسائل اربط Google من الملف الشخصي.",
        "Demande congé": "طلب عطلة",
        "Choisis le type, la période et le destinataire. Les week-ends et jours fériés ne sont pas comptés.": "اختر النوع والفترة والمستلم. لا يتم احتساب عطلة نهاية الأسبوع والعطل الرسمية.",
        "Occasions": "المناسبات",
        "Pour naissance, mariage ou décès, renseigne les jours autorisés non déduits. Le reste est calculé automatiquement.": "للولادة أو الزواج أو الوفاة، أدخل الأيام المرخصة غير المخصومة. يتم حساب الباقي تلقائياً.",
        "Maladie": "المرض",
        "Pour repos maladie, renseigne une référence/certificat. Évite de saisir des données médicales sensibles inutiles.": "للراحة المرضية، أدخل مرجعاً/شهادة. تجنب إدخال بيانات طبية حساسة غير ضرورية.",
        "Google Calendar": "تقويم Google",
        "Un congé approuvé peut être synchronisé dans Google Calendar. En cas d’annulation, l’événement est supprimé.": "يمكن مزامنة العطلة الموافق عليها مع تقويم Google. عند الإلغاء يتم حذف الحدث.",
        "Exports": "التصدير",
        "Tu peux télécharger les bilans en Excel et PDF depuis le menu.": "يمكنك تنزيل التقارير Excel و PDF من القائمة.",
        "Abonnement": "الاشتراك",
        "Les utilisateurs non-admin doivent avoir un abonnement actif pour accéder aux fonctions principales.": "يحتاج المستخدمون غير المديرين إلى اشتراك نشط للوصول للوظائف الرئيسية.",
        "Sécurité": "الأمان",
        "Utilise un mot de passe fort, garde ton compte Google sécurisé et ne partage jamais tes accès.": "استخدم كلمة مرور قوية، حافظ على أمان حساب Google ولا تشارك بيانات الدخول.",
        "Données séparées": "بيانات منفصلة",
        "Les données de Mohamed AIT ELMALEM appartiennent au compte owner. Chaque nouvel utilisateur commence à zéro et ajoute ses propres congés.": "بيانات المالك تخص حساب المالك فقط. كل مستخدم جديد يبدأ من الصفر ويضيف عطله الخاصة.",
        "Nouvelle demande de congé": "طلب عطلة جديد",
        "Portail collaborateur": "بوابة الموظف",
        "Choisis la période avec calendrier. Les week-ends et jours fériés ne sont pas comptés.": "اختر الفترة عبر التقويم. لا يتم احتساب نهاية الأسبوع والعطل الرسمية.",
        "Créer la demande": "إنشاء الطلب",
        "Dernières demandes": "آخر الطلبات",
        "Synchroniser avec calendrier": "مزامنة مع التقويم",
        "Toutes les demandes": "كل الطلبات",
        "Synchronisation depuis Google Calendar": "مزامنة من تقويم Google",
        "Cette page affiche toutes les demandes et peut importer les congés déjà présents dans Google Calendar.": "تعرض هذه الصفحة كل الطلبات ويمكنها استيراد العطل الموجودة في تقويم Google.",
        "Import congés existants": "استيراد عطل موجودة",
        "Depuis": "من",
        "Jusqu’à": "إلى",
        "Calendriers Maroc : Standard & Hijri": "تقويمات المغرب: عادي وهجري",
        "Sélectionne le calendrier à synchroniser depuis Google Calendar : jours fériés standards ou événements Hijri.": "اختر التقويم للمزامنة من Google Calendar: العطل الرسمية أو الأحداث الهجرية.",
        "Profil": "الملف الشخصي",
        "État Google": "حالة Google",
        "Google non connecté. Connecte-toi avec Google pour envoyer les emails depuis ton adresse.": "Google غير متصل. اربط Google لإرسال الرسائل من بريدك.",
        "Connecter Google": "ربط Google",
        "Enregistrer": "حفظ",
        "Backups base de données": "نسخ قاعدة البيانات",
        "Protection des données": "حماية البيانات",
        "Crée, télécharge et restaure des sauvegardes SQLite. Accès réservé admin.": "إنشاء وتنزيل واستعادة نسخ SQLite. للمدير فقط.",
        "Stratégie recommandée": "الاستراتيجية المقترحة",
        "Créer un backup maintenant": "إنشاء نسخة الآن",
        "Fichiers backup": "ملفات النسخ الاحتياطي",
        "Uploader un backup": "رفع نسخة احتياطية",
        "Restore sécurisé": "استعادة آمنة",
        "Seuls les fichiers SQLite sont autorisés.": "يسمح فقط بملفات SQLite.",
        "Historique des demandes": "سجل الطلبات",
        "Traçabilité complète": "تتبع كامل",
        "Les congés déjà pris sont chargés ici comme demandes approuvées, avec période, jours pris, statut et commentaire.": "العطل السابقة تُعرض هنا كطلبات موافق عليها مع الفترة والأيام والحالة والتعليق.",
        "Liste des demandes": "قائمة الطلبات",
        "Demandes & Google Calendar": "الطلبات و Google Calendar",
        "Annuler": "إلغاء",
        "Motif annulation": "سبب الإلغاء",
        "Synchronisé depuis le bilan fourni": "تمت المزامنة من الحصيلة المقدمة",
        "Congé historique approuvé": "عطلة تاريخية موافق عليها",
        "Congé annuel payé": "عطلة سنوية مدفوعة",
        "Profiling": "تحليل الأداء",
        "Performances": "الأداء"
    }
}


def tr_text(text):
    try:
        lang = get_lang()
    except Exception:
        lang = "fr"
    if not text:
            pass
        return text
    if lang == "fr":
            pass
        return text
    return I18N_TEXT_MAP.get(lang, {}).get(str(text).strip(), text)


# -------------------------
# MEDFLOW V35 FULL LANGUAGE PACK
# -------------------------
V35_I18N = {
    "fr": {
        "app_name":"MEDFLOW","public_home":"Accueil","home_title":"Gestion professionnelle des congés, simplement.","home_subtitle":"MEDFLOW centralise les congés, validations, calendriers, exports et abonnements dans un espace premium.",
        "start_trial":"Commencer l’essai 7 jours","sign_in":"Se connecter","create_account":"Créer un compte","login_title":"Connexion sécurisée","login_subtitle":"Connecte-toi à ton espace sécurisé MEDFLOW.",
        "username":"Nom utilisateur","password":"Mot de passe","google_login":"Connexion Google","discover_platform":"Découvrir la plateforme","already_registered":"Déjà inscrit ? Se connecter",
        "register_title":"Créer un compte","register_subtitle":"Crée ton compte avec 7 jours d’essai gratuits. Tes données démarrent à zéro.","full_name":"Nom complet","email":"Email","company":"Entreprise","job_title":"Poste","hire_date":"Date d’embauche","confirm_password":"Confirmer mot de passe",
        "dashboard":"Dashboard","guide":"Guide utilisateur","subscription":"Abonnement","leave_request":"Demande congé","history":"Historique","calendar":"Calendrier","holidays":"Jours fériés","profile":"Profil","admin_db":"Admin DB","backups":"Backups","profiling":"Profiling","export_excel":"Export Excel","export_pdf":"Export PDF","logout":"Déconnexion",
        "language":"Langue","mode":"Mode","dark":"Dark","light":"Light","owner":"Owner","active":"Actif","inactive":"Inactif","trial":"Essai","workspace":"Workspace","version":"Version",
        "about":"À propos","privacy":"Confidentialité","terms":"Conditions","copyright":"Droits d’auteur","about_title":"À propos de nous","about_desc":"MEDFLOW est une plateforme moderne pour gérer les congés, absences, calendriers et abonnements de manière simple et sécurisée.","our_mission":"Notre mission","mission_desc":"Offrir un outil clair, accessible et sécurisé pour les utilisateurs individuels, freelances et petites équipes.",
        "privacy_title":"Politique de confidentialité","privacy_desc":"Les données sont utilisées uniquement pour gérer les congés, les utilisateurs, les abonnements et les intégrations Google autorisées.","terms_title":"Conditions d’utilisation","terms_desc":"L’utilisateur est responsable de l’exactitude des demandes saisies. MEDFLOW aide à gérer les congés mais ne remplace pas un conseil juridique officiel.","copyright_title":"Droits d’auteur","copyright_desc":"© MEDFLOW. Tous droits réservés. Le nom, le logo, le design et le code sont protégés.",
        "pricing_title":"Abonnement MEDFLOW","secure_payment":"Paiement sécurisé","trial_7":"7 jours d’essai","monthly":"Mensuel","annual":"Annuel","use_trial":"Utiliser l’essai","subscribe_monthly":"S’abonner mensuellement","subscribe_annually":"S’abonner annuellement","manage_billing":"Gérer facturation","payment_setup":"Configuration paiement owner","payment_setup_desc":"Pour recevoir les paiements, crée deux prix Stripe puis ajoute les variables dans Render.",
        "dashboard_title":"Tableau de bord","dashboard_subtitle":"Gestion congés secteur privé Maroc","monthly_summary":"Bilan mensuel","leave_balance":"Solde congé","approved_leaves":"Congés approuvés","pending_requests":"Demandes en attente","rejected_requests":"Demandes refusées","taken_leaves":"Congés pris","end_balance":"Solde fin",
        "year":"Année","all":"Tous","month":"Mois","credit":"Crédit","extra":"Extra","taken":"Pris","period":"Période","balance":"Solde","status":"Statut","actions":"Actions","created_at":"Créée le","type":"Type","days":"Jours","comment":"Commentaire","decision":"Décision","source":"Source","date":"Date","name":"Nom",
        "new_leave_request":"Nouvelle demande de congé","request_intro":"Choisis la période avec calendrier. Les week-ends et jours fériés ne sont pas comptés.","leave_type":"Type de congé","start_date":"Date début","end_date":"Date fin","recipient_email":"Email destinataire","email_message":"Message email personnalisé","send_email":"Envoyer email via Gmail API OAuth","create_request":"Créer la demande","authorized_days":"Jours autorisés occasion / loi","estimated_working_days":"Jours ouvrables estimés","deduct_from_balance":"À déduire du solde","medical_certificate":"Certificat médical sécurisé","latest_requests":"Dernières demandes","direct_tracking":"Suivi direct",
        "requests_list":"Liste des demandes","requests_google":"Demandes & Google Calendar","sync_from_google":"Synchronisation depuis Google Calendar","sync_calendar":"Synchroniser avec calendrier","import_existing":"Import congés existants","from":"Depuis","to":"Jusqu’à","all_requests":"Toutes les demandes","google_calendar":"Google Calendar","synced":"Synchronisé","not_synced":"Non","cancel":"Annuler","cancel_reason":"Motif annulation",
        "guide_title":"Guide utilisateur MEDFLOW","guide_subtitle":"Tout ce qu’il faut savoir pour utiliser l’application simplement et en sécurité.","section_login":"Connexion","section_leave":"Demande congé","section_occasions":"Occasions","section_sick":"Maladie","section_calendar":"Google Calendar","section_exports":"Exports","section_subscription":"Abonnement","section_security":"Sécurité","section_data":"Données séparées",
        "profile_title":"Profil","google_status":"État Google","google_connected":"Google connecté","google_not_connected":"Google non connecté","connect_google":"Connecter Google","disconnect_google":"Déconnecter Google","save":"Enregistrer",
        "holidays_title":"Calendriers Maroc : Standard & Hijri","holidays_subtitle":"Sélectionne le calendrier à synchroniser depuis Google Calendar.","standard":"Standard","hijri":"Hijri",
        "backups_title":"Backups base de données","data_protection":"Protection des données","create_backup":"Créer un backup maintenant","backup_files":"Fichiers backup","upload_backup":"Uploader un backup","restore_secure":"Restore sécurisé","download":"Télécharger","restore":"Restaurer","delete":"Supprimer","edit":"Modifier","admin_only":"Admin uniquement","no_data":"Aucune donnée",
        "approved":"Approuvé","pending":"En attente","refused":"Refusé","cancelled":"Annulé",
        "annual":"Congé annuel payé","birth":"Naissance fils/fille","death_parent":"Décès parent / ascendant","death_spouse_child":"Décès conjoint / enfant","death_sibling_parent_inlaw":"Décès frère/sœur/beau-parent","marriage_self":"Mariage du salarié","marriage_child":"Mariage d’un enfant","circumcision":"Circoncision","operation_family":"Opération conjoint/enfant à charge","sick":"Repos maladie",
        "january":"Janvier","february":"Février","march":"Mars","april":"Avril","may":"Mai","june":"Juin","july":"Juillet","august":"Août","september":"Septembre","october":"Octobre","november":"Novembre","december":"Décembre"
    },
    "en": {
        "app_name":"MEDFLOW","public_home":"Home","home_title":"Professional leave management, beautifully simple.","home_subtitle":"MEDFLOW centralizes leave, approvals, calendars, exports and subscriptions in a premium workspace.",
        "start_trial":"Start 7-day trial","sign_in":"Sign in","create_account":"Create account","login_title":"Secure login","login_subtitle":"Sign in to your secure MEDFLOW workspace.",
        "username":"Username","password":"Password","google_login":"Google login","discover_platform":"Discover platform","already_registered":"Already registered? Sign in",
        "register_title":"Create account","register_subtitle":"Create your account with 7 free trial days. Your data starts from zero.","full_name":"Full name","email":"Email","company":"Company","job_title":"Job title","hire_date":"Hire date","confirm_password":"Confirm password",
        "dashboard":"Dashboard","guide":"User guide","subscription":"Subscription","leave_request":"Leave request","history":"History","calendar":"Calendar","holidays":"Holidays","profile":"Profile","admin_db":"Admin DB","backups":"Backups","profiling":"Profiling","export_excel":"Excel export","export_pdf":"PDF export","logout":"Logout",
        "language":"Language","mode":"Mode","dark":"Dark","light":"Light","owner":"Owner","active":"Active","inactive":"Inactive","trial":"Trial","workspace":"Workspace","version":"Version",
        "about":"About","privacy":"Privacy","terms":"Terms","copyright":"Copyright","about_title":"About us","about_desc":"MEDFLOW is a modern platform for managing leave, absences, calendars and subscriptions simply and securely.","our_mission":"Our mission","mission_desc":"Provide a clear, accessible and secure tool for individuals, freelancers and small teams.",
        "privacy_title":"Privacy policy","privacy_desc":"Data is used only to manage leave, users, subscriptions and authorized Google integrations.","terms_title":"Terms of use","terms_desc":"The user is responsible for the accuracy of submitted requests. MEDFLOW helps manage leave but does not replace official legal advice.","copyright_title":"Copyright","copyright_desc":"© MEDFLOW. All rights reserved. The name, logo, design and code are protected.",
        "pricing_title":"MEDFLOW subscription","secure_payment":"Secure payment","trial_7":"7-day trial","monthly":"Monthly","annual":"Annual","use_trial":"Use trial","subscribe_monthly":"Subscribe monthly","subscribe_annually":"Subscribe annually","manage_billing":"Manage billing","payment_setup":"Owner payment setup","payment_setup_desc":"To receive payments, create two Stripe prices then add the variables in Render.",
        "dashboard_title":"Dashboard","dashboard_subtitle":"Morocco private sector leave management","monthly_summary":"Monthly summary","leave_balance":"Leave balance","approved_leaves":"Approved leaves","pending_requests":"Pending requests","rejected_requests":"Rejected requests","taken_leaves":"Taken leaves","end_balance":"End balance",
        "year":"Year","all":"All","month":"Month","credit":"Credit","extra":"Extra","taken":"Taken","period":"Period","balance":"Balance","status":"Status","actions":"Actions","created_at":"Created at","type":"Type","days":"Days","comment":"Comment","decision":"Decision","source":"Source","date":"Date","name":"Name",
        "new_leave_request":"New leave request","request_intro":"Choose the period with the calendar. Weekends and holidays are not counted.","leave_type":"Leave type","start_date":"Start date","end_date":"End date","recipient_email":"Recipient email","email_message":"Custom email message","send_email":"Send email via Gmail API OAuth","create_request":"Create request","authorized_days":"Authorized occasion / legal days","estimated_working_days":"Estimated working days","deduct_from_balance":"To deduct from balance","medical_certificate":"Secure medical certificate","latest_requests":"Latest requests","direct_tracking":"Direct tracking",
        "requests_list":"Requests list","requests_google":"Requests & Google Calendar","sync_from_google":"Sync from Google Calendar","sync_calendar":"Sync with calendar","import_existing":"Import existing leaves","from":"From","to":"To","all_requests":"All requests","google_calendar":"Google Calendar","synced":"Synced","not_synced":"No","cancel":"Cancel","cancel_reason":"Cancellation reason",
        "guide_title":"MEDFLOW user guide","guide_subtitle":"Everything you need to use the app simply and securely.","section_login":"Login","section_leave":"Leave request","section_occasions":"Occasions","section_sick":"Sickness","section_calendar":"Google Calendar","section_exports":"Exports","section_subscription":"Subscription","section_security":"Security","section_data":"Separated data",
        "profile_title":"Profile","google_status":"Google status","google_connected":"Google connected","google_not_connected":"Google not connected","connect_google":"Connect Google","disconnect_google":"Disconnect Google","save":"Save",
        "holidays_title":"Morocco calendars: Standard & Hijri","holidays_subtitle":"Select the calendar to synchronize from Google Calendar.","standard":"Standard","hijri":"Hijri",
        "backups_title":"Database backups","data_protection":"Data protection","create_backup":"Create backup now","backup_files":"Backup files","upload_backup":"Upload backup","restore_secure":"Secure restore","download":"Download","restore":"Restore","delete":"Delete","edit":"Edit","admin_only":"Admin only","no_data":"No data",
        "approved":"Approved","pending":"Pending","refused":"Rejected","cancelled":"Cancelled",
        "annual":"Paid annual leave","birth":"Birth of child","death_parent":"Death of parent / ascendant","death_spouse_child":"Death of spouse / child","death_sibling_parent_inlaw":"Death of sibling / parent-in-law","marriage_self":"Employee marriage","marriage_child":"Child marriage","circumcision":"Circumcision","operation_family":"Spouse/dependent child operation","sick":"Sick leave",
        "january":"January","february":"February","march":"March","april":"April","may":"May","june":"June","july":"July","august":"August","september":"September","october":"October","november":"November","december":"December"
    }
}
# Generate ES/DE/AR by overriding important keys
V35_I18N["de"] = dict(V35_I18N["en"], **{
"public_home":"Startseite","home_title":"Professionelle Urlaubsverwaltung, schön einfach.","home_subtitle":"MEDFLOW zentralisiert Urlaub, Genehmigungen, Kalender, Exporte und Abonnements in einem Premium-Arbeitsbereich.",
"start_trial":"7 Tage testen","sign_in":"Anmelden","create_account":"Konto erstellen","login_title":"Sichere Anmeldung","login_subtitle":"Melden Sie sich in Ihrem sicheren MEDFLOW-Bereich an.","username":"Benutzername","password":"Passwort","register_title":"Konto erstellen","full_name":"Vollständiger Name","company":"Unternehmen","job_title":"Position","hire_date":"Einstellungsdatum","confirm_password":"Passwort bestätigen",
"guide":"Benutzerhandbuch","subscription":"Abonnement","leave_request":"Urlaubsantrag","history":"Verlauf","calendar":"Kalender","holidays":"Feiertage","profile":"Profil","logout":"Abmelden","language":"Sprache","active":"Aktiv","inactive":"Inaktiv","trial":"Test","about":"Über uns","privacy":"Datenschutz","terms":"Bedingungen","copyright":"Urheberrecht","about_title":"Über uns","our_mission":"Unsere Mission",
"pricing_title":"MEDFLOW Abonnement","secure_payment":"Sichere Zahlung","trial_7":"7 Tage Testphase","monthly":"Monatlich","annual":"Jährlich","use_trial":"Testphase nutzen","subscribe_monthly":"Monatlich abonnieren","subscribe_annually":"Jährlich abonnieren","manage_billing":"Abrechnung verwalten",
"dashboard_title":"Dashboard","dashboard_subtitle":"Urlaubsverwaltung Privatsektor Marokko","monthly_summary":"Monatsübersicht","leave_balance":"Urlaubssaldo","approved_leaves":"Genehmigte Urlaube","pending_requests":"Ausstehende Anträge","rejected_requests":"Abgelehnte Anträge","taken_leaves":"Genommene Urlaube","end_balance":"Endsaldo",
"year":"Jahr","all":"Alle","month":"Monat","credit":"Guthaben","taken":"Genommen","period":"Zeitraum","balance":"Saldo","status":"Status","actions":"Aktionen","created_at":"Erstellt am","type":"Typ","days":"Tage","comment":"Kommentar","decision":"Entscheidung","date":"Datum","name":"Name",
"new_leave_request":"Neuer Urlaubsantrag","request_intro":"Wählen Sie den Zeitraum im Kalender. Wochenenden und Feiertage werden nicht gezählt.","leave_type":"Urlaubsart","start_date":"Startdatum","end_date":"Enddatum","recipient_email":"Empfänger-E-Mail","email_message":"Individuelle E-Mail-Nachricht","create_request":"Antrag erstellen","estimated_working_days":"Geschätzte Arbeitstage","deduct_from_balance":"Vom Saldo abzuziehen","medical_certificate":"Sicheres ärztliches Attest","latest_requests":"Letzte Anträge",
"requests_list":"Antragsliste","sync_from_google":"Synchronisierung aus Google Calendar","sync_calendar":"Mit Kalender synchronisieren","from":"Von","to":"Bis","all_requests":"Alle Anträge","synced":"Synchronisiert","cancel":"Stornieren","cancel_reason":"Stornierungsgrund",
"profile_title":"Profil","google_status":"Google-Status","connect_google":"Google verbinden","save":"Speichern","backups_title":"Datenbank-Backups","create_backup":"Backup jetzt erstellen","backup_files":"Backup-Dateien","upload_backup":"Backup hochladen","download":"Herunterladen","restore":"Wiederherstellen","delete":"Löschen","edit":"Bearbeiten",
"approved":"Genehmigt","pending":"Ausstehend","refused":"Abgelehnt","cancelled":"Storniert","annual":"Bezahlter Jahresurlaub","birth":"Geburt eines Kindes","sick":"Krankschreibung"
})
V35_I18N["es"] = dict(V35_I18N["en"], **{
"public_home":"Inicio","home_title":"Gestión profesional de permisos, simple y elegante.","start_trial":"Prueba de 7 días","sign_in":"Iniciar sesión","create_account":"Crear cuenta","login_title":"Inicio seguro","username":"Usuario","password":"Contraseña","register_title":"Crear cuenta","full_name":"Nombre completo","company":"Empresa","job_title":"Puesto","hire_date":"Fecha de contratación","confirm_password":"Confirmar contraseña","guide":"Guía de usuario","subscription":"Suscripción","leave_request":"Solicitud de permiso","history":"Historial","calendar":"Calendario","holidays":"Festivos","profile":"Perfil","logout":"Cerrar sesión","language":"Idioma","active":"Activo","inactive":"Inactivo","trial":"Prueba","about":"Sobre nosotros","privacy":"Privacidad","terms":"Condiciones","copyright":"Derechos de autor","pricing_title":"Suscripción MEDFLOW","secure_payment":"Pago seguro","trial_7":"Prueba de 7 días","monthly":"Mensual","annual":"Anual","dashboard_title":"Panel","year":"Año","all":"Todos","month":"Mes","period":"Periodo","balance":"Saldo","status":"Estado","actions":"Acciones","new_leave_request":"Nueva solicitud de permiso","start_date":"Fecha inicio","end_date":"Fecha fin","recipient_email":"Email destinatario","create_request":"Crear solicitud","approved":"Aprobado","pending":"Pendiente","refused":"Rechazado","cancelled":"Cancelado","annual":"Vacaciones anuales pagadas","sick":"Baja médica"
})
V35_I18N["ar"] = dict(V35_I18N["en"], **{
"public_home":"الرئيسية","home_title":"تدبير احترافي للعطل ببساطة وجمال.","home_subtitle":"MEDFLOW يجمع العطل والموافقات والتقويمات والتصدير والاشتراكات في مساحة عمل احترافية.","start_trial":"ابدأ تجربة 7 أيام","sign_in":"تسجيل الدخول","create_account":"إنشاء حساب","login_title":"تسجيل دخول آمن","login_subtitle":"سجّل الدخول إلى فضاء MEDFLOW الآمن.","username":"اسم المستخدم","password":"كلمة المرور","register_title":"إنشاء حساب","register_subtitle":"أنشئ حسابك مع 7 أيام تجربة مجانية. تبدأ بياناتك من الصفر.","full_name":"الاسم الكامل","email":"البريد الإلكتروني","company":"الشركة","job_title":"المنصب","hire_date":"تاريخ التوظيف","confirm_password":"تأكيد كلمة المرور",
"dashboard":"لوحة التحكم","guide":"دليل المستخدم","subscription":"الاشتراك","leave_request":"طلب عطلة","history":"السجل","calendar":"التقويم","holidays":"العطل الرسمية","profile":"الملف الشخصي","logout":"تسجيل الخروج","language":"اللغة","mode":"الوضع","active":"نشط","inactive":"غير نشط","trial":"تجربة","workspace":"مساحة العمل",
"about":"من نحن","privacy":"الخصوصية","terms":"الشروط","copyright":"حقوق النشر","about_title":"من نحن","about_desc":"MEDFLOW منصة حديثة لتدبير العطل والغيابات والتقويمات والاشتراكات بطريقة بسيطة وآمنة.","our_mission":"مهمتنا","mission_desc":"تقديم أداة واضحة وآمنة وسهلة الاستخدام للأفراد والمستقلين والفرق الصغيرة.",
"privacy_title":"سياسة الخصوصية","terms_title":"شروط الاستخدام","copyright_title":"حقوق النشر","copyright_desc":"© MEDFLOW. جميع الحقوق محفوظة.",
"pricing_title":"اشتراك MEDFLOW","secure_payment":"دفع آمن","trial_7":"تجربة 7 أيام","monthly":"شهري","annual":"سنوي","use_trial":"استخدم التجربة","subscribe_monthly":"اشترك شهرياً","subscribe_annually":"اشترك سنوياً","manage_billing":"إدارة الفوترة",
"dashboard_title":"لوحة التحكم","dashboard_subtitle":"تدبير العطل للقطاع الخاص بالمغرب","monthly_summary":"الحصيلة الشهرية","leave_balance":"رصيد العطل","approved_leaves":"العطل الموافق عليها","pending_requests":"طلبات قيد الانتظار","rejected_requests":"طلبات مرفوضة","taken_leaves":"عطل مأخوذة","end_balance":"الرصيد النهائي",
"year":"السنة","all":"الكل","month":"الشهر","credit":"الرصيد","extra":"إضافي","taken":"المأخوذ","period":"الفترة","balance":"الرصيد","status":"الحالة","actions":"الإجراءات","created_at":"تاريخ الإنشاء","type":"النوع","days":"الأيام","comment":"تعليق","decision":"القرار","source":"المصدر","date":"التاريخ","name":"الاسم",
"new_leave_request":"طلب عطلة جديد","request_intro":"اختر الفترة عبر التقويم. لا يتم احتساب نهاية الأسبوع والعطل الرسمية.","leave_type":"نوع العطلة","start_date":"تاريخ البداية","end_date":"تاريخ النهاية","recipient_email":"بريد المستلم","email_message":"رسالة بريد مخصصة","send_email":"إرسال عبر Gmail API OAuth","create_request":"إنشاء الطلب","authorized_days":"الأيام المرخصة / القانونية","estimated_working_days":"أيام العمل المقدرة","deduct_from_balance":"سيتم خصمها من الرصيد","medical_certificate":"شهادة طبية آمنة","latest_requests":"آخر الطلبات","direct_tracking":"تتبع مباشر",
"requests_list":"قائمة الطلبات","requests_google":"الطلبات و Google Calendar","sync_from_google":"مزامنة من Google Calendar","sync_calendar":"مزامنة مع التقويم","import_existing":"استيراد عطل موجودة","from":"من","to":"إلى","all_requests":"كل الطلبات","google_calendar":"تقويم Google","synced":"متزامن","not_synced":"لا","cancel":"إلغاء","cancel_reason":"سبب الإلغاء",
"profile_title":"الملف الشخصي","google_status":"حالة Google","google_connected":"Google متصل","google_not_connected":"Google غير متصل","connect_google":"ربط Google","disconnect_google":"فصل Google","save":"حفظ","backups_title":"نسخ قاعدة البيانات","data_protection":"حماية البيانات","create_backup":"إنشاء نسخة الآن","backup_files":"ملفات النسخ الاحتياطي","upload_backup":"رفع نسخة احتياطية","restore_secure":"استعادة آمنة","download":"تحميل","restore":"استعادة","delete":"حذف","edit":"تعديل","admin_only":"للمدير فقط","no_data":"لا توجد بيانات",
"approved":"موافق عليه","pending":"قيد الانتظار","refused":"مرفوض","cancelled":"ملغى","annual":"عطلة سنوية مدفوعة","birth":"ازدياد ابن/ابنة","death_parent":"وفاة أحد الوالدين / الأصول","death_spouse_child":"وفاة الزوج/الزوجة أو الابن","marriage_self":"زواج الموظف","marriage_child":"زواج ابن/ابنة","sick":"راحة مرضية"
})

def t35(key):
    try:
        lang = get_lang()
    except Exception:
        lang = "fr"
    return V35_I18N.get(lang, V35_I18N["fr"]).get(key, key)

def t35_text(text):
    if not text:
            pass
        return text
    try:
        lang = get_lang()
    except Exception:
        lang = "fr"
    if lang == "fr":
            pass
        return text
    # reverse search by FR text then map to target
    for k, v in V35_I18N["fr"].items():
        if str(v).strip() == str(text).strip():
                pass
            return V35_I18N.get(lang, {}).get(k, text)
    return text

def month_name_i18n_v35(m):
    keys = ["january","february","march","april","may","june","july","august","september","october","november","december"]
    return t35(keys[int(m)-1])

def status_i18n_v35(s):
    return t35(str(s or "").lower())

def leave_label_i18n_v35(key):
    return t35(str(key or "annual"))

def current_user():
    uid = session.get("uid")
    return db.session.get(User, uid) if uid else None

def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user():
                pass
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


def is_admin_user(u):
    return bool(u and u.email and u.email.lower() == ADMIN_EMAIL)

def has_active_subscription(u):
    return bool(is_admin_user(u) or (u and u.subscription_status in ["active", "trialing"]))

def subscription_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        u = current_user()
        if not u:
                pass
            return redirect(url_for("login"))
        if not has_active_subscription(u):
                pass
            flash("Abonnement requis pour accéder à cette fonctionnalité.", "warning")
            return redirect(url_for("pricing"))
        return fn(*args, **kwargs)
    return wrapper


@app.before_request
def medflow_register_template_globals():
    try:
        app.jinja_env.globals["tr_text"] = tr_text
        app.jinja_env.globals["tt"] = tt if "tt" in globals() else tr_text
    except Exception:
        pass


@app.before_request
def enforce_admin_email_role():
    u = current_user()
    if u and u.email and u.email.lower() == ADMIN_EMAIL and u.role != "admin":
        u.role = "admin"
        u.subscription_status = "active"
        db.session.commit()


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        u = current_user()
        if not u or u.email.lower() != ADMIN_EMAIL:
            abort(403)
        return fn(*args, **kwargs)
    return wrapper

def parse_date(s):
    if not s:
        raise ValueError("Date obligatoire.")
    return datetime.strptime(s, "%Y-%m-%d").date()

def month_credit(user, y, m):
    # 18 jours/an = 1,5/mois au départ. +1,5/an chaque 5 ans, plafonné à 30j/an.
    years = y - user.hire_date.year - ((m, 1) < (user.hire_date.month, user.hire_date.day))
    extra_yearly = max(0, years // 5) * 1.5
    yearly = min(30, 18 + extra_yearly)
    return round(yearly / 12, 2)

def is_weekend(d): return d.weekday() in (5,6)
def holiday_set():
    return {h.day for h in Holiday.query.all()}
def working_days_between(start, end):
    hols = holiday_set()
    n = 0
    days = []
    d = start
    while d <= end:
        if not is_weekend(d) and d not in hols:
            n += 1; days.append(d)
        d += timedelta(days=1)
    return n, days

def period_label(start,end):
    return start.strftime("%d/%m/%Y") if start == end else f"{start.strftime('%d/%m/%Y')}-{end.strftime('%d/%m/%Y')}"

def add_or_update_month(y,m,credit,extra,taken,period,balance,locked=True):
    row = MonthlyBalance.query.filter_by(year=y, month=m).first()
    if not row:
        row = MonthlyBalance(year=y, month=m)
        db.session.add(row)
    row.credit, row.extra, row.taken, row.period, row.balance, row.locked_seed = credit, extra, taken, period, balance, locked

def seed_holidays_to(year_to=2100):
    for y in range(2023, year_to+1):
        for (m,d), name in FIXED_MA_HOLIDAYS.items():
            dt = date(y,m,d)
            if not Holiday.query.filter_by(day=dt).first():
                    pass
                db.session.add(Holiday(day=dt, name=name, source="system", hijri=False))
        if y in HIJRI_FALLBACK:
            for ds, name in HIJRI_FALLBACK[y]:
                dt = datetime.strptime(ds, "%Y-%m-%d").date()
                if not Holiday.query.filter_by(day=dt).first():
                        pass
                    db.session.add(Holiday(day=dt, name=name, source="fallback", hijri=True))
    db.session.commit()

def seed_db():
    """V36: démarrage propre.
    - Aucun congé historique préchargé.
    - Aucun bilan Mohamed préchargé.
    - Création uniquement du compte owner/admin si absent.
    """
    db.create_all()
    owner_email = os.getenv("ADMIN_EMAIL", "aitelmalemmohamed@gmail.com").lower()
    admin = User.query.filter_by(email=owner_email).first()
    if not admin:
        admin = User(
            username="admin",
            email=owner_email,
            password_hash=generate_password_hash(os.getenv("ADMIN_PASSWORD", "ChangeMe-Admin-2026!")),
            full_name="Owner",
            company="MEDFLOW",
            job_title="Owner",
            hire_date=date.today(),
            role="admin",
            subscription_status="active",
            trial_ends_at=datetime.utcnow()+timedelta(days=36500),
            initial_balance=0,
            initial_balance_date=date.today()
        )
        db.session.add(admin)
        db.session.commit()

def recalc_from_seed(user, from_year=2028, to_year=None):
    if to_year is None: to_year = max(date.today().year+5, 2030)
    prev = MonthlyBalance.query.filter_by(year=2027,month=12).first()
    balance = prev.balance if prev else 0
    for y in range(from_year, to_year+1):
        for m in range(1,13):
            approved = LeaveRequest.query.filter_by(status="approved", leave_type="annual").all()
            taken = 0
            periods = []
            for r in approved:
                for d in dates_in_range(r.start_date, r.end_date):
                    if d.year==y and d.month==m and not is_weekend(d) and d not in holiday_set():
                        taken += 1
                if r.start_date.year==y and r.start_date.month==m:
                    periods.append(period_label(r.start_date,r.end_date))
            credit = month_credit(user,y,m)
            extra = 0
            balance = round(balance + credit + extra - taken, 2)
            add_or_update_month(y,m,credit,extra,taken,", ".join(periods),balance,False)
    db.session.commit()

def dates_in_range(start,end):
    d=start
    while d<=end:
        yield d
        d+=timedelta(days=1)

def get_google_creds():
    u=current_user()
    if not u or not u.google_token: return None
    data=json.loads(u.google_token)
    return Credentials.from_authorized_user_info(data, SCOPES)

def save_google_creds(creds):
    u=current_user()
    u.google_token = creds.to_json()
    db.session.commit()


def gregorian_to_hijri_approx(gdate):
    """Approximation Hijri pour affichage UI. Les jours fériés officiels restent synchronisés via Google Calendar."""
    import math
    y, m, d = gdate.year, gdate.month, gdate.day
    if m < 3:
        y -= 1
        m += 12
    a = math.floor(y / 100)
    b = 2 - a + math.floor(a / 4)
    jd = math.floor(365.25 * (y + 4716)) + math.floor(30.6001 * (m + 1)) + d + b - 1524
    islamic = jd - 1948440 + 10632
    n = math.floor((islamic - 1) / 10631)
    islamic = islamic - 10631 * n + 354
    j = (math.floor((10985 - islamic) / 5316)) * (math.floor((50 * islamic) / 17719)) + (math.floor(islamic / 5670)) * (math.floor((43 * islamic) / 15238))
    islamic = islamic - (math.floor((30 - j) / 15)) * (math.floor((17719 * j) / 50)) - (math.floor(j / 16)) * (math.floor((15238 * j) / 43)) + 29
    hm = math.floor((24 * islamic) / 709)
    hd = islamic - math.floor((709 * hm) / 24)
    hy = 30 * n + j - 30
    months = ["محرم","صفر","ربيع 1","ربيع 2","جمادى 1","جمادى 2","رجب","شعبان","رمضان","شوال","ذو القعدة","ذو الحجة"]
    hm = max(1, min(12, hm))
    return f"{hd} {months[hm-1]} {hy}H"



@app.after_request
def add_security_headers(response):
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"] = "no-referrer"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    response.headers["Cache-Control"] = "no-store"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "img-src 'self' data: https:; "
        "style-src 'self' 'unsafe-inline' https:; "
        "script-src 'self' 'unsafe-inline' https://js.stripe.com https://accounts.google.com; "
        "frame-src https://js.stripe.com https://accounts.google.com; "
        "connect-src 'self' https://www.googleapis.com https://oauth2.googleapis.com;"
    )
    return response


@app.before_request
def ensure_i18n_globals():
    app.jinja_env.globals["t"] = safe_t
    app.jinja_env.globals["languages"] = LANGUAGES if "LANGUAGES" in globals() else {"fr": "Français"}
    app.jinja_env.globals["current_lang"] = get_lang() if "get_lang" in globals() else "fr"
    app.jinja_env.globals["is_rtl"] = (get_lang() == "ar") if "get_lang" in globals() else False

@app.context_processor
def inject():
    return {"user": current_user(), "months_fr": MONTHS_FR, "leave_types": LEAVE_TYPES, "hijri_date": gregorian_to_hijri_approx, "theme": (current_user().theme if current_user() else "dark"), "ADMIN_EMAIL": ADMIN_EMAIL, "has_active_subscription": has_active_subscription, "is_admin_user": is_admin_user}

@app.route("/")
@login_required
@subscription_required
def dashboard():
    u=current_user()
    year=int(request.args.get("year", date.today().year))
    # V36: no historical recalc on startup.year+2))
    rows=MonthlyBalance.query.filter_by(year=year).order_by(MonthlyBalance.month).all()
    pending_q=LeaveRequest.query.filter_by(status="pending")
    approved_q=LeaveRequest.query.filter_by(status="approved")
    refused_q=LeaveRequest.query.filter_by(status="refused")
    if u.role != "admin":
        pending_q = pending_q.filter_by(user_id=u.id)
        approved_q = approved_q.filter_by(user_id=u.id)
        refused_q = refused_q.filter_by(user_id=u.id)
    pending=pending_q.order_by(LeaveRequest.created_at.desc()).all()
    approved=approved_q.order_by(LeaveRequest.created_at.desc()).limit(8).all()
    refused=refused_q.count()
    total_taken=sum([r.taken for r in rows]) if rows else 0
    end_balance=rows[-1].balance if rows else 0
    years=list(range(2023, date.today().year + 2))
    return render_template("dashboard.html", rows=rows, year=year, years=years, pending=pending, approved=approved, refused=refused, total_taken=total_taken, end_balance=end_balance)



@app.route("/home")
def public_home():
    return render_template("public_home.html")

@app.route("/register", methods=["GET","POST"])
def register():
    if request.method == "POST":
        try:
            hire_raw = request.form.get("hire_date")
            if hire_raw:
                u.hire_date = datetime.strptime(hire_raw, "%Y-%m-%d").date()

            u.initial_balance = float(request.form.get("initial_balance") or 0)

            raw_date = request.form.get("initial_balance_date")
            if raw_date:
                u.initial_balance_date = datetime.strptime(raw_date, "%Y-%m-%d").date()
            else:
                u.initial_balance_date = date.today()

            db.session.commit()
            flash("Compte créé avec succès. Tu peux te connecter.", "success")
            return redirect(url_for("login"))
        except Exception as e:
            db.session.rollback()
            flash(str(e), "danger")
    return render_template("register.html")


@app.route("/login", methods=["GET","POST"])
def login():
    if request.method=="POST":
        username=request.form["username"].strip()
        password=request.form["password"]
        u=User.query.filter((User.username==username)|(User.email==username)).first()
        if u and u.password_hash and check_password_hash(u.password_hash,password):
            session["uid"]=u.id
            return redirect(url_for("dashboard"))
        flash("Identifiant ou mot de passe incorrect.", "danger")
    return render_template("login.html")


@app.route("/toggle_theme")
@login_required
def toggle_theme():
    u = current_user()
    u.theme = "light" if (u.theme or "dark") == "dark" else "dark"
    db.session.commit()
    return redirect(request.referrer or url_for("dashboard"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/google_login")
def google_login():
    cid=os.getenv("GOOGLE_CLIENT_ID"); sec=os.getenv("GOOGLE_CLIENT_SECRET")
    if not cid or not sec:
            pass
        flash("Google OAuth non configuré dans .env", "danger")
        return redirect(url_for("login"))
    flow=Flow.from_client_config({"web":{
        "client_id":cid,"client_secret":sec,
        "auth_uri":"https://accounts.google.com/o/oauth2/auth",
        "token_uri":"https://oauth2.googleapis.com/token",
        "redirect_uris":[os.getenv("GOOGLE_REDIRECT_URI","http://127.0.0.1:5000/oauth2callback")]
    }}, scopes=SCOPES)
    flow.redirect_uri=os.getenv("GOOGLE_REDIRECT_URI","http://127.0.0.1:5000/oauth2callback")

    # PKCE : Google peut demander un code_verifier au callback.
    # On le génère au login et on le garde en session jusqu'au callback.
    code_verifier = secrets.token_urlsafe(64)
    code_challenge = base64.urlsafe_b64encode(
        hashlib.sha256(code_verifier.encode("utf-8")).digest()
    ).decode("utf-8").rstrip("=")

    auth_url, state = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
        code_challenge=code_challenge,
        code_challenge_method="S256"
    )
    session["oauth_state"] = state
    session["code_verifier"] = code_verifier
    return redirect(auth_url)

@app.route("/oauth2callback")
def oauth2callback():
    cid=os.getenv("GOOGLE_CLIENT_ID"); sec=os.getenv("GOOGLE_CLIENT_SECRET")
    flow=Flow.from_client_config({"web":{
        "client_id":cid,"client_secret":sec,
        "auth_uri":"https://accounts.google.com/o/oauth2/auth",
        "token_uri":"https://oauth2.googleapis.com/token",
        "redirect_uris":[os.getenv("GOOGLE_REDIRECT_URI","http://127.0.0.1:5000/oauth2callback")]
    }}, scopes=SCOPES, state=session.get("oauth_state"))
    flow.redirect_uri=os.getenv("GOOGLE_REDIRECT_URI","http://127.0.0.1:5000/oauth2callback")
    try:
        code_verifier = session.get("code_verifier")
        if not code_verifier:
            raise Exception("Session OAuth expirée : code_verifier manquant. Recommence la connexion Google.")
        flow.fetch_token(
            authorization_response=request.url,
            code_verifier=code_verifier
        )
        creds=flow.credentials

        svc=build("oauth2","v2",credentials=creds)
        info=svc.userinfo().get().execute()
        email=info.get("email")
        if not email:
            raise Exception("Google n'a pas retourné l'email du compte.")

        u=User.query.filter_by(email=email).first()
        if not u:
            u=User(email=email, username=email.split("@")[0], full_name=info.get("name","Utilisateur Google"))
            db.session.add(u)
            db.session.commit()

        session["uid"]=u.id
        session.pop("code_verifier", None)
        save_google_creds(creds)
        flash("Connexion Google réussie.", "success")
        return redirect(url_for("dashboard"))
    except Exception as exc:
        db.session.rollback()
        flash("Erreur OAuth Google : " + str(exc), "danger")
        return redirect(url_for("login"))


@app.route("/disconnect_google", methods=["POST"])
@login_required
def disconnect_google():
    u = current_user()
    u.google_token = None
    db.session.commit()
    flash("Compte Google déconnecté.", "success")
    return redirect(url_for("profile"))

@app.route("/profile", methods=["GET","POST"])
@login_required
@subscription_required
def profile():
    u=current_user()
    if request.method=="POST":
        u.full_name=request.form["full_name"]
        u.company=request.form["company"]
        u.job_title=request.form["job_title"]
        u.hire_date=parse_date(request.form["hire_date"])
        db.session.commit()
        flash("Profil mis à jour.", "success")
        return redirect(url_for("profile"))
    return render_template("profile.html")

@app.route("/request_leave", methods=["GET","POST"])
@login_required
@subscription_required
def request_leave():
    if request.method=="POST":
        try:
            leave_type=request.form["leave_type"]
            start=parse_date(request.form["start_date"]); end=parse_date(request.form["end_date"])
            if start>end: raise ValueError("La date début doit être inférieure ou égale à la date fin.")
            workdays, days = working_days_between(start,end)
            if workdays<=0: raise ValueError("La période choisie ne contient aucun jour ouvrable calculable.")
            maxdays=LEAVE_TYPES[leave_type]["default"]
            if maxdays and workdays>maxdays:
                raise ValueError(f"Ce type de congé autorise normalement {maxdays} jours.")
            if leave_type == "sick" and not request.form.get("medical_note"):
                raise ValueError("Certificat médical / référence obligatoire pour un repos maladie.")
            overlap=LeaveRequest.query.filter(LeaveRequest.status.in_(["pending","approved"]), LeaveRequest.start_date<=end, LeaveRequest.end_date>=start).first()
            if overlap:
                raise ValueError("Cette période chevauche déjà une demande en attente ou approuvée.")
            lr=LeaveRequest(user_id=current_user().id, leave_type=leave_type, start_date=start, end_date=end, working_days=workdays, recipient=request.form.get("recipient"), message=request.form.get("message"), medical_note=request.form.get("medical_note"))
            db.session.add(lr); db.session.commit()
            if request.form.get("send_email"):
                send_gmail(lr)
                flash("Demande créée et email envoyé via Gmail API.", "success")
            else:
                flash("Demande créée en attente d'approbation.", "success")
            return redirect(url_for("history"))
        except Exception as e:
            db.session.rollback(); flash(str(e), "danger")
    recent_q = LeaveRequest.query
    if current_user().role != "admin":
        recent_q = recent_q.filter_by(user_id=current_user().id)
    recent_requests = recent_q.order_by(LeaveRequest.created_at.desc()).limit(12).all()
    return render_template("request_leave.html", recent_requests=recent_requests)

def send_gmail(lr):
    if not lr.recipient:
        raise Exception("Aucun destinataire email renseigné.")
    creds=get_google_creds()
    if not creds:
        raise Exception("Connecte-toi avec Google pour envoyer via ton adresse Gmail.")
    try:
        service=build("gmail","v1",credentials=creds)
        u=current_user()
        body=render_template("email_leave.txt", lr=lr, u=u)
        msg=MIMEText(body, "plain", "utf-8")
        msg["to"]=lr.recipient
        msg["subject"]=f"Demande de congé - {LEAVE_TYPES[lr.leave_type]['label']} - {u.full_name}"
        raw=base64.urlsafe_b64encode(msg.as_bytes()).decode()
        service.users().messages().send(userId="me", body={"raw":raw}).execute()
    except Exception as exc:
        raise Exception(f"Erreur d'envoi Gmail API avec le compte utilisateur : {exc}")
    except Exception as exc:
        raise Exception(f"Erreur d'envoi Gmail API : {exc}")


@app.route("/cancel/<int:rid>", methods=["POST"])
@login_required
@admin_required
def cancel_leave(rid):
    lr=db.session.get(LeaveRequest,rid) or abort(404)
    if lr.status not in ["pending", "approved"]:
            pass
        flash("Cette demande ne peut pas être annulée.", "warning")
        return redirect(request.referrer or url_for("history"))

    old_status = lr.status
    lr.status = "cancelled"
    lr.decided_at = datetime.utcnow()
    reason = request.form.get("comment") or "Annulation pour urgence / cas particulier"
    lr.decision_comment = reason

    # Si un événement Google Calendar avait été créé, on le supprime.
    if lr.calendar_event_id:
        try:
            delete_google_calendar_event(lr.calendar_event_id)
            lr.calendar_event_id = None
            flash("Congé annulé et événement Google Calendar supprimé.", "success")
        except Exception as e:
            flash("Congé annulé, mais suppression Google Calendar échouée: " + str(e), "warning")
    else:
        flash("Congé annulé.", "success")

    # Si le congé était approuvé et impactait le solde, on recalcule.
    if old_status == "approved" and (lr.deductible_days or 0) > 0:
    # V36: no historical recalc on startup, 2028, max(lr.end_date.year, date.today().year+2))

    db.session.commit()
    return redirect(request.referrer or url_for("history"))

def delete_google_calendar_event(event_id):
    creds=get_google_creds()
    if not creds:
        raise Exception("Connexion Google requise pour supprimer l'événement Calendar.")
    svc=build("calendar","v3",credentials=creds)
    svc.events().delete(calendarId="primary", eventId=event_id).execute()



@app.route("/requests")
@login_required
@subscription_required
def requests_list():
    year = request.args.get("year", "all")
    q = LeaveRequest.query
    if current_user().role != "admin":
        q = q.filter(LeaveRequest.user_id == current_user().id)
    if year != "all":
        q = q.filter(db.extract("year", LeaveRequest.start_date) == int(year))
    rows = q.order_by(LeaveRequest.created_at.desc()).all()
    years = list(range(2023, date.today().year + 2))
    return render_template("requests.html", rows=rows, year=year, years=years)

@app.route("/sync_calendar_leaves", methods=["POST"])
@login_required
@admin_required
def sync_calendar_leaves():
    y1 = int(request.form.get("from_year", date.today().year))
    y2 = int(request.form.get("to_year", date.today().year))
    try:
        count = import_leave_events_from_google_calendar(y1, y2)
        flash(f"{count} congé(s) importé(s) depuis Google Calendar.", "success")
    except Exception as e:
        flash("Synchronisation des congés depuis Google Calendar échouée : " + str(e), "danger")
    return redirect(url_for("requests_list", year=y1))

def import_leave_events_from_google_calendar(y1, y2):
    creds = get_google_creds()
    if not creds:
        raise Exception("Connexion Google requise.")
    svc = build("calendar", "v3", credentials=creds)
    events = svc.events().list(
        calendarId="primary",
        timeMin=f"{y1}-01-01T00:00:00Z",
        timeMax=f"{y2+1}-01-01T00:00:00Z",
        singleEvents=True,
        maxResults=2500,
        q="Congé"
    ).execute().get("items", [])

    imported = 0
    u = current_user()
    for ev in events:
        start_raw = ev.get("start", {}).get("date") or ev.get("start", {}).get("dateTime", "")[:10]
        end_raw = ev.get("end", {}).get("date") or ev.get("end", {}).get("dateTime", "")[:10]
        if not start_raw or not end_raw:
            continue

        start = datetime.strptime(start_raw, "%Y-%m-%d").date()
        end_exclusive = datetime.strptime(end_raw, "%Y-%m-%d").date()
        end = end_exclusive - timedelta(days=1) if end_exclusive > start else start

        summary = ev.get("summary", "Congé Google Calendar")
        event_id = ev.get("id")

        # Eviter les doublons : par event_id ou même période approuvée.
        exists = LeaveRequest.query.filter(
            (LeaveRequest.calendar_event_id == event_id) |
            ((LeaveRequest.start_date == start) & (LeaveRequest.end_date == end) & (LeaveRequest.status.in_(["approved","pending"])))
        ).first()
        if exists:
            continue

        workdays, _ = working_days_between(start, end)
        if workdays <= 0:
            continue

        lr = LeaveRequest(
            user_id=u.id,
            leave_type="annual",
            start_date=start,
            end_date=end,
            working_days=workdays,
            exceptional_paid_days=0,
            deductible_days=workdays,
            status="approved",
            message="Importé automatiquement depuis Google Calendar : " + summary,
            decision_comment="Synchronisé depuis Google Calendar",
            decided_at=datetime.utcnow(),
            calendar_event_id=event_id
        )
        db.session.add(lr)
        imported += 1

    db.session.commit()
    if imported:
    # V36: no historical recalc on startup.year + 2))
    return imported


@app.route("/history")
@login_required
@subscription_required
def history():
    year = request.args.get("year", "all")
    q = LeaveRequest.query
    if current_user().role != "admin":
        q = q.filter(LeaveRequest.user_id == current_user().id)
    if year != "all":
        q = q.filter(db.extract("year", LeaveRequest.start_date) == int(year))
    rows=q.order_by(LeaveRequest.created_at.desc()).all()
    years=list(range(2023, date.today().year + 2))
    return render_template("history.html", rows=rows, year=year, years=years)

@app.route("/approve/<int:rid>", methods=["POST"])
@login_required
@admin_required
def approve(rid):
    lr=db.session.get(LeaveRequest,rid) or abort(404)
    if lr.status!="pending":
            pass
        flash("Demande déjà traitée.", "warning"); return redirect(url_for("history"))
    lr.status="approved"; lr.decided_at=datetime.utcnow(); lr.decision_comment=request.form.get("comment")
    if LEAVE_TYPES[lr.leave_type]["deduct"]:
        apply_leave_to_balances(lr)
    if request.form.get("calendar"):
        try: create_google_calendar_event(lr)
        except Exception as e: flash("Congé approuvé, mais import Google Calendar échoué: "+str(e), "warning")
    db.session.commit()
    flash("Demande approuvée.", "success")
    return redirect(url_for("history"))

@app.route("/refuse/<int:rid>", methods=["POST"])
@login_required
@admin_required
def refuse(rid):
    lr=db.session.get(LeaveRequest,rid) or abort(404)
    lr.status="refused"; lr.decided_at=datetime.utcnow(); lr.decision_comment=request.form.get("comment")
    db.session.commit()
    flash("Demande refusée. Aucun solde n'a été modifié.", "info")
    return redirect(url_for("history"))

def apply_leave_to_balances(lr):
    u=db.session.get(User, lr.user_id)
    y0=min(lr.start_date.year, 2028)
    # V36: no historical recalc on startup.year+2))

def create_google_calendar_event(lr):
    creds=get_google_creds()
    if not creds: raise Exception("Connexion Google requise.")
    svc=build("calendar","v3",credentials=creds)
    ev={
        "summary": f"Congé - {LEAVE_TYPES[lr.leave_type]['label']}",
        "description": lr.message or "",
        "start":{"date":lr.start_date.isoformat()},
        "end":{"date":(lr.end_date+timedelta(days=1)).isoformat()},
    }
    res=svc.events().insert(calendarId="primary", body=ev).execute()
    lr.calendar_event_id=res.get("id")

@app.route("/calendar")
@login_required
@subscription_required
def calendar_view():
    year=int(request.args.get("year", date.today().year))
    approved_q=LeaveRequest.query.filter_by(status="approved")
    if current_user().role != "admin":
        approved_q = approved_q.filter_by(user_id=current_user().id)
    approved=approved_q.all()
    xdays=set()
    for lr in approved:
        for d in dates_in_range(lr.start_date,lr.end_date):
            xdays.add(d)
    hols={h.day:h.name for h in Holiday.query.filter(Holiday.day>=date(year,1,1), Holiday.day<=date(year,12,31)).all()}
    years=list(range(2023, date.today().year + 2))
    return render_template("calendar.html", year=year, years=years, xdays=xdays, hols=hols, calendar=calendar, date=date)

@app.route("/sync_holidays", methods=["POST"])
@login_required
@admin_required
def sync_holidays():
    y1=int(request.form.get("from_year", 2023))
    y2=int(request.form.get("to_year", date.today().year+1))
    calendar_type=request.form.get("calendar_type", "standard")
    try:
        count = sync_google_holidays(y1, y2, calendar_type)
        label = "Standard Maroc" if calendar_type == "standard" else "Hijri Maroc"
        flash(f"{count} jours fériés synchronisés depuis Google Calendar ({label}).", "success")
    except Exception as e:
        flash("Sync Google Calendar impossible, fallback local conservé: "+str(e), "warning")
    return redirect(url_for("holidays", year=y1, calendar_type=calendar_type))

def sync_google_holidays(y1, y2, calendar_type="standard"):
    creds=get_google_creds()
    if not creds:
        raise Exception("connecte-toi avec Google.")

    if calendar_type == "hijri":
        cal_id=os.getenv("GOOGLE_HIJRI_CALENDAR_ID","ar.ma#holiday@group.v.calendar.google.com")
        force_hijri=True
        source_name="google_hijri"
    else:
        cal_id=os.getenv("GOOGLE_HOLIDAY_CALENDAR_ID","fr.ma#holiday@group.v.calendar.google.com")
        force_hijri=False
        source_name="google_standard"

    svc=build("calendar","v3",credentials=creds)
    events=svc.events().list(
        calendarId=cal_id,
        timeMin=f"{y1}-01-01T00:00:00Z",
        timeMax=f"{y2+1}-01-01T00:00:00Z",
        singleEvents=True,
        maxResults=2500
    ).execute().get("items",[])

    count = 0
    for e in events:
        ds=e["start"].get("date")
        if ds:
            dt=datetime.strptime(ds,"%Y-%m-%d").date()
            name=e.get("summary","Jour férié")
            hijri=force_hijri or any(k.lower() in name.lower() for k in ["aïd","aid","fitr","adha","moharram","mawlid","hégire","hijri","ramadan"])
            h=Holiday.query.filter_by(day=dt).first()
            if h:
                # On garde le nom le plus précis; Hijri prioritaire si importé.
                h.name=name
                h.source=source_name
                h.hijri=hijri
            else:
                db.session.add(Holiday(day=dt,name=name,source=source_name,hijri=hijri))
            count += 1

    db.session.commit()
    return count

@app.route("/holidays")
@login_required
@subscription_required
def holidays():
    year=int(request.args.get("year", date.today().year))
    calendar_type=request.args.get("calendar_type", "all")
    q=Holiday.query.filter(Holiday.day>=date(year,1,1), Holiday.day<=date(year,12,31))
    if calendar_type == "hijri":
        q=q.filter(Holiday.hijri == True)
    elif calendar_type == "standard":
        q=q.filter(Holiday.hijri == False)
    rows=q.order_by(Holiday.day).all()
    years=list(range(2023, date.today().year + 2))
    return render_template("holidays.html", rows=rows, year=year, years=years, calendar_type=calendar_type)

@app.route("/export_excel")
@login_required
@subscription_required
def export_excel():
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "MEDFLOW Bilan"

    u = current_user()
    ws.merge_cells("A1:G1")
    ws["A1"] = "Bilan des congés - MEDFLOW"
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1E293B")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"{u.full_name} | {u.job_title} | Embauche : {u.hire_date.strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(size=11, italic=True, color="334155")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Année","Mois","Congé / Mois","Jours supplémentaires","Congé pris","Période","Solde restant"]
    ws.append([])
    ws.append(headers)
    header_row = 4

    dark = PatternFill("solid", fgColor="0F172A")
    blue = PatternFill("solid", fgColor="DBEAFE")
    green = PatternFill("solid", fgColor="DCFCE7")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")

    for cell in ws[header_row]:
        cell.fill = dark
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for r in MonthlyBalance.query.order_by(MonthlyBalance.year, MonthlyBalance.month):
        ws.append([r.year, MONTHS_FR[r.month-1], r.credit, r.extra, r.taken, r.period, r.balance])

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")
        row[6].fill = green if (row[6].value or 0) >= 0 else PatternFill("solid", fgColor="FEE2E2")
        row[1].fill = blue

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = [12,16,16,20,14,30,16][col-1]
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "38BDF8"

    path=os.path.join(app.instance_path,"bilan_conges_medflow.xlsx")
    wb.save(path)
    return send_file(path, as_attachment=True)

@app.route("/export_pdf")
@login_required
@subscription_required
def export_pdf():
    path=os.path.join(app.instance_path,"bilan_conges_medflow.pdf")
    doc=SimpleDocTemplate(path,pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles=getSampleStyleSheet()
    u = current_user()

    title = Paragraph("<b>Bilan des congés - MEDFLOW</b>", styles["Title"])
    subtitle = Paragraph(f"{u.full_name} | {u.job_title} | Date d'embauche : {u.hire_date.strftime('%d/%m/%Y')}", styles["Normal"])

    data=[["Année","Mois","Crédit","Extra","Pris","Période","Solde"]]
    for r in MonthlyBalance.query.order_by(MonthlyBalance.year, MonthlyBalance.month):
        data.append([r.year, MONTHS_FR[r.month-1], r.credit, r.extra, r.taken, r.period, r.balance])

    t=Table(data, repeatRows=1, colWidths=[55,90,70,70,60,230,70])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0F172A")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("ALIGN",(0,0),(-1,0),"CENTER"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.HexColor("#F8FAFC"), colors.HexColor("#E0F2FE")]),
        ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#CBD5E1")),
        ("FONT",(0,0),(-1,-1),"Helvetica",8),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TEXTCOLOR",(6,1),(6,-1),colors.HexColor("#0369A1")),
    ]))
    doc.build([title, Spacer(1,6), subtitle, Spacer(1,16), t])
    return send_file(path, as_attachment=True)



@app.route("/admin/change_password", methods=["GET", "POST"])
@login_required
@admin_required
def admin_change_password():
    u = current_user()
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")
        if not u.password_hash or not check_password_hash(u.password_hash, current):
                pass
            flash("Mot de passe actuel incorrect.", "danger")
        elif len(new) < 14:
                pass
            flash("Le nouveau mot de passe doit contenir au moins 14 caractères.", "danger")
        elif new != confirm:
                pass
            flash("Confirmation différente.", "danger")
        else:
            u.password_hash = generate_password_hash(new)
            db.session.commit()
            flash("Mot de passe admin modifié.", "success")
            return redirect(url_for("dashboard"))
    return render_template("admin_change_password.html")



@app.route("/guide")
@login_required
def guide():
    return render_template("guide.html")

@app.route("/privacy")
def privacy():
    return render_template("privacy.html")

@app.route("/terms")
def terms():
    return render_template("terms.html")

@app.route("/pricing")
@login_required
def pricing():
    return render_template("pricing.html", stripe_public_key=os.getenv("STRIPE_PUBLISHABLE_KEY", ""), price_id=os.getenv("STRIPE_PRICE_ID", ""))

@app.route("/create_checkout_session", methods=["POST"])
@login_required
def create_checkout_session():
    u = current_user()
    if is_admin_user(u):
            pass
        flash("Le compte admin a un accès complet.", "info")
        return redirect(url_for("dashboard"))
    if not os.getenv("STRIPE_SECRET_KEY") or not os.getenv("STRIPE_PRICE_ID"):
            pass
        flash("Paiement non configuré : ajoute STRIPE_SECRET_KEY et STRIPE_PRICE_ID dans Render.", "danger")
        return redirect(url_for("pricing"))
    domain = os.getenv("APP_BASE_URL", request.url_root.rstrip("/"))
    try:
        checkout_session = stripe.checkout.Session.create(
            mode="subscription",
            payment_method_types=["card"],
            line_items=[{"price": os.getenv("STRIPE_PRICE_ID"), "quantity": 1}],
            customer_email=u.email,
            client_reference_id=str(u.id),
            billing_address_collection="auto",
            allow_promotion_codes=True,
            metadata={"user_id": str(u.id), "app": "MEDFLOW"},
            subscription_data={"metadata": {"user_id": str(u.id), "app": "MEDFLOW"}},
            success_url=domain + url_for("payment_success") + "?session_id={CHECKOUT_SESSION_ID}",
            cancel_url=domain + url_for("payment_cancel"),
        )
        return redirect(checkout_session.url, code=303)
    except Exception as e:
        flash("Erreur paiement : " + str(e), "danger")
        return redirect(url_for("pricing"))


@app.route("/billing_portal", methods=["POST"])
@login_required
def billing_portal():
    u = current_user()
    if is_admin_user(u):
            pass
        flash("Le compte owner n’a pas besoin d’espace facturation.", "info")
        return redirect(url_for("dashboard"))
    if not os.getenv("STRIPE_SECRET_KEY"):
            pass
        flash("Stripe non configuré.", "danger")
        return redirect(url_for("pricing"))
    if not u.stripe_customer_id:
            pass
        flash("Aucun client Stripe lié à ce compte. Lance d’abord un abonnement.", "warning")
        return redirect(url_for("pricing"))
    try:
        domain = os.getenv("APP_BASE_URL", request.url_root.rstrip("/"))
        session_portal = stripe.billing_portal.Session.create(
            customer=u.stripe_customer_id,
            return_url=domain + url_for("pricing"),
        )
        return redirect(session_portal.url, code=303)
    except Exception as e:
        flash("Erreur portail facturation : " + str(e), "danger")
        return redirect(url_for("pricing"))

@app.route("/payment_success")
@login_required
def payment_success():
    session_id = request.args.get("session_id")
    u = current_user()
    try:
        if session_id and os.getenv("STRIPE_SECRET_KEY"):
            checkout_session = stripe.checkout.Session.retrieve(session_id)
            if str(checkout_session.get("client_reference_id")) == str(u.id) and checkout_session.get("payment_status") in ["paid", "no_payment_required"]:
                u.subscription_status = "active"
                u.stripe_customer_id = checkout_session.get("customer")
                u.stripe_subscription_id = checkout_session.get("subscription")
                db.session.commit()
                flash("Paiement confirmé. Abonnement activé.", "success")
                return redirect(url_for("dashboard"))
    except Exception:
        pass
    flash("Paiement reçu. Ton abonnement sera activé après confirmation Stripe.", "success")
    return redirect(url_for("dashboard"))

@app.route("/payment_cancel")
@login_required
def payment_cancel():
    flash("Paiement annulé.", "warning")
    return redirect(url_for("pricing"))

@app.route("/stripe_webhook", methods=["POST"])
def stripe_webhook():
    payload = request.get_data(as_text=True)
    sig_header = request.headers.get("Stripe-Signature")
    endpoint_secret = os.getenv("STRIPE_WEBHOOK_SECRET", "")
    try:
        if endpoint_secret:
            event = stripe.Webhook.construct_event(payload, sig_header, endpoint_secret)
        else:
            event = stripe.Event.construct_from(request.get_json(force=True), stripe.api_key)
    except Exception:
        return "Invalid payload", 400

    event_type = event["type"]
    data = event["data"]["object"]

    if event_type == "checkout.session.completed":
        user_id = data.get("client_reference_id")
        u = db.session.get(User, int(user_id)) if user_id else None
        if u:
            u.subscription_status = "active"
            u.stripe_customer_id = data.get("customer")
            u.stripe_subscription_id = data.get("subscription")
            db.session.commit()

    elif event_type in ["customer.subscription.deleted", "customer.subscription.paused"]:
        sub_id = data.get("id")
        u = User.query.filter_by(stripe_subscription_id=sub_id).first()
        if u and not is_admin_user(u):
            u.subscription_status = "inactive"
            db.session.commit()

    elif event_type in ["customer.subscription.updated"]:
        sub_id = data.get("id")
        status = data.get("status")
        u = User.query.filter_by(stripe_subscription_id=sub_id).first()
        if u and not is_admin_user(u):
            u.subscription_status = status or "inactive"
            db.session.commit()

    return "ok", 200



# -------------------------
# Backup / Restore SQLite
# -------------------------
def db_file_path():
    return os.path.join(app.instance_path, "conges_medflow.db")

def create_db_backup(reason="manual"):
    src = db_file_path()
    if not os.path.exists(src):
        raise Exception("Base SQLite introuvable.")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_reason = "".join(ch for ch in reason if ch.isalnum() or ch in ["_", "-"])[:30] or "backup"
    dst = os.path.join(BACKUP_DIR, f"medflow_{safe_reason}_{stamp}.db")
    shutil.copy2(src, dst)
    cleanup_old_backups()
    return dst

def cleanup_old_backups(keep=20):
    files = sorted(glob.glob(os.path.join(BACKUP_DIR, "*.db")), key=os.path.getmtime, reverse=True)
    for old in files[keep:]:
        try:
            os.remove(old)
        except Exception:
            pass

def auto_backup_once_per_day():
    try:
        today = datetime.now().strftime("%Y%m%d")
        marker = os.path.join(BACKUP_DIR, f".auto_{today}")
        if os.path.exists(db_file_path()) and not os.path.exists(marker):
            create_db_backup("auto")
            with open(marker, "w", encoding="utf-8") as f:
                f.write(datetime.now().isoformat())
    except Exception:
        pass

@app.route("/admin/backups")
@login_required
@admin_required
def admin_backups():
    files = []
    for p in sorted(glob.glob(os.path.join(BACKUP_DIR, "*.db")), key=os.path.getmtime, reverse=True):
        files.append({
            "name": os.path.basename(p),
            "size": round(os.path.getsize(p) / 1024, 2),
            "date": datetime.fromtimestamp(os.path.getmtime(p)).strftime("%d/%m/%Y %H:%M:%S")
        })
    return render_template("admin_backups.html", files=files)

@app.route("/admin/backups/create", methods=["POST"])
@login_required
@admin_required
def admin_backup_create():
    try:
        create_db_backup("manual")
        flash("Backup créé avec succès.", "success")
    except Exception as e:
        flash("Erreur backup : " + str(e), "danger")
    return redirect(url_for("admin_backups"))

@app.route("/admin/backups/download/<filename>")
@login_required
@admin_required
def admin_backup_download(filename):
    filename = secure_filename(filename)
    path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(path):
        abort(404)
    return send_file(path, as_attachment=True)

@app.route("/admin/backups/restore", methods=["POST"])
@login_required
@admin_required
def admin_backup_restore():
    filename = secure_filename(request.form.get("filename", ""))
    confirm = request.form.get("confirm", "")
    if confirm != "RESTORE":
            pass
        flash("Restauration annulée : écris RESTORE pour confirmer.", "danger")
        return redirect(url_for("admin_backups"))
    backup_path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(backup_path):
            pass
        flash("Backup introuvable.", "danger")
        return redirect(url_for("admin_backups"))
    try:
        create_db_backup("before_restore")
        db.session.remove()
        shutil.copy2(backup_path, db_file_path())
        flash("Base restaurée avec succès. Redémarre le service Render si nécessaire.", "success")
    except Exception as e:
        flash("Erreur restauration : " + str(e), "danger")
    return redirect(url_for("admin_backups"))


# Mini panneau Admin SQLite
ADMIN_DB_MODELS = {
    "users": User,
    "leave_requests": LeaveRequest,
    "monthly_balances": MonthlyBalance,
    "holidays": Holiday,
}

@app.route("/admin/db")
@login_required
@admin_required
def admin_db():
    tables = list(ADMIN_DB_MODELS.keys())
    table = request.args.get("table", "leave_requests")
    if table not in ADMIN_DB_MODELS:
        table = "leave_requests"
    model = ADMIN_DB_MODELS[table]
    rows = model.query.order_by(model.id.desc()).limit(300).all()
    columns = [c.name for c in model.__table__.columns if c.name not in ["password_hash", "google_token", "stripe_customer_id", "stripe_subscription_id"]]
    return render_template("admin_db.html", tables=tables, table=table, rows=rows, columns=columns)

@app.route("/admin/db/edit/<table>/<int:rid>", methods=["GET", "POST"])
@login_required
@admin_required
def admin_db_edit(table, rid):
    if table not in ADMIN_DB_MODELS:
        abort(404)
    model = ADMIN_DB_MODELS[table]
    row = db.session.get(model, rid) or abort(404)
    columns = [c for c in model.__table__.columns if c.name not in ["id", "password_hash", "google_token"]]
    if request.method == "POST":
        try:
            for col in columns:
                name = col.name
                raw = request.form.get(name)
                current = getattr(row, name)
                if raw == "":
                    value = None
                elif isinstance(current, bool):
                    value = raw in ["1", "true", "True", "on"]
                elif hasattr(current, "year") and hasattr(current, "month") and hasattr(current, "day") and not hasattr(current, "hour"):
                    value = datetime.strptime(raw, "%Y-%m-%d").date()
                elif hasattr(current, "hour") and hasattr(current, "minute"):
                    value = datetime.strptime(raw, "%Y-%m-%dT%H:%M")
                elif isinstance(current, int):
                    value = int(raw)
                elif isinstance(current, float):
                    value = float(raw)
                else:
                    value = raw
                setattr(row, name, value)
            db.session.commit()
            flash("Ligne modifiée avec succès.", "success")
            return redirect(url_for("admin_db", table=table))
        except Exception as e:
            db.session.rollback()
            flash("Erreur modification : " + str(e), "danger")
    return render_template("admin_db_edit.html", table=table, row=row, columns=columns)

@app.route("/admin/db/delete/<table>/<int:rid>", methods=["POST"])
@login_required
@admin_required
def admin_db_delete(table, rid):
    if table not in ADMIN_DB_MODELS:
        abort(404)
    model = ADMIN_DB_MODELS[table]
    row = db.session.get(model, rid) or abort(404)
    try:
        db.session.delete(row)
        db.session.commit()
        flash("Ligne supprimée.", "success")
    except Exception as e:
        db.session.rollback()
        flash("Erreur suppression : " + str(e), "danger")
    return redirect(url_for("admin_db", table=table))


@app.route("/about")
def about():
    return render_template("about.html")

@app.route("/copyright")
def copyright_page():
    return render_template("copyright.html")



# -------------------------
# Admin Profiling
# -------------------------
PROFILE_REQUESTS = deque(maxlen=500)
PROFILE_STATS = defaultdict(lambda: {"count": 0, "total": 0.0, "max": 0.0})

@app.before_request
def medflow_profile_start():
    request._mf_start_time = time.perf_counter()

@app.after_request
def medflow_profile_end(response):
    try:
        elapsed = (time.perf_counter() - getattr(request, "_mf_start_time", time.perf_counter())) * 1000
        endpoint = request.endpoint or request.path
        PROFILE_STATS[endpoint]["count"] += 1
        PROFILE_STATS[endpoint]["total"] += elapsed
        PROFILE_STATS[endpoint]["max"] = max(PROFILE_STATS[endpoint]["max"], elapsed)
        PROFILE_REQUESTS.appendleft({
            "path": request.path,
            "endpoint": endpoint,
            "method": request.method,
            "status": response.status_code,
            "ms": round(elapsed, 2),
            "time": datetime.now().strftime("%H:%M:%S")
        })
    except Exception:
        pass
    return response

@app.route("/admin/profiling")
@login_required
@admin_required
def admin_profiling():
    stats = []
    for endpoint, s in PROFILE_STATS.items():
        avg = s["total"] / s["count"] if s["count"] else 0
        stats.append({
            "endpoint": endpoint,
            "count": s["count"],
            "avg": round(avg, 2),
            "max": round(s["max"], 2),
            "total": round(s["total"], 2)
        })
    stats = sorted(stats, key=lambda x: x["avg"], reverse=True)
    return render_template("admin_profiling.html", stats=stats, requests=list(PROFILE_REQUESTS))

@app.errorhandler(403)
def forbidden(e):
    return render_template("error.html", message="Accès refusé."), 403

@app.errorhandler(404)
def not_found(e):
    return render_template("error.html", message="Page introuvable."), 404

@app.errorhandler(Exception)
def handle_error(e):
    db.session.rollback()
    return render_template("error.html", message=str(e)), 500


# Initialisation globale Jinja pour éviter UndefinedError sur pages d'erreur
try:
    app.jinja_env.globals["t"] = safe_t
except Exception:
    pass


try:
    app.jinja_env.globals["tt"] = tt
    app.jinja_env.globals["month_name_i18n"] = month_name_i18n
    app.jinja_env.globals["status_i18n"] = status_i18n
    app.jinja_env.globals["leave_label_i18n"] = leave_label_i18n
except Exception:
    pass


# V34: Jinja globals safety
try:
    app.jinja_env.globals["tr_text"] = tr_text
    app.jinja_env.globals["tt"] = tt if "tt" in globals() else tr_text
except Exception:
    pass




def patch_v36_i18n():
    extra = {
        "fr": {
            "initial_setup":"Configuration initiale",
            "initial_setup_desc":"Renseigne ta date d’embauche et le solde de congé que tu possèdes actuellement. MEDFLOW démarre ensuite à partir de ces valeurs.",
            "initial_balance":"Solde initial",
            "initial_balance_date":"Date du solde initial",
            "start_from_zero":"Le site démarre à zéro",
            "no_seed_data":"Aucune donnée utilisateur n’est préchargée.",
            "setup_now":"Configurer maintenant",
            "clean_start":"Démarrage propre"
        },
        "en": {
            "initial_setup":"Initial setup",
            "initial_setup_desc":"Enter your hire date and the leave balance you currently have. MEDFLOW will start from these values.",
            "initial_balance":"Initial balance",
            "initial_balance_date":"Initial balance date",
            "start_from_zero":"The site starts from zero",
            "no_seed_data":"No user data is preloaded.",
            "setup_now":"Set up now",
            "clean_start":"Clean start"
        },
        "de": {
            "initial_setup":"Ersteinrichtung",
            "initial_setup_desc":"Geben Sie Ihr Einstellungsdatum und den aktuellen Urlaubssaldo ein. MEDFLOW startet mit diesen Werten.",
            "initial_balance":"Anfangssaldo",
            "initial_balance_date":"Datum des Anfangssaldos",
            "start_from_zero":"Die Website startet bei null",
            "no_seed_data":"Es sind keine Benutzerdaten vorgeladen.",
            "setup_now":"Jetzt einrichten",
            "clean_start":"Sauberer Start"
        },
        "es": {
            "initial_setup":"Configuración inicial",
            "initial_setup_desc":"Introduce tu fecha de contratación y el saldo de permisos actual. MEDFLOW empezará desde esos valores.",
            "initial_balance":"Saldo inicial",
            "initial_balance_date":"Fecha del saldo inicial",
            "start_from_zero":"El sitio empieza desde cero",
            "no_seed_data":"No hay datos de usuario precargados.",
            "setup_now":"Configurar ahora",
            "clean_start":"Inicio limpio"
        },
        "ar": {
            "initial_setup":"الإعداد الأولي",
            "initial_setup_desc":"أدخل تاريخ التوظيف ورصيد العطل الحالي. سيبدأ MEDFLOW انطلاقاً من هذه القيم.",
            "initial_balance":"الرصيد الأولي",
            "initial_balance_date":"تاريخ الرصيد الأولي",
            "start_from_zero":"الموقع يبدأ من الصفر",
            "no_seed_data":"لا توجد بيانات مستخدمين محملة مسبقاً.",
            "setup_now":"الإعداد الآن",
            "clean_start":"بداية نظيفة"
        }
    }
    try:
        for lang, vals in extra.items():
            V35_I18N.setdefault(lang, {})
            V35_I18N[lang].update(vals)
    except Exception:
        pass
patch_v36_i18n()

# V35 template globals
try:
    app.jinja_env.globals["t35"] = t35
    app.jinja_env.globals["tt"] = t35
    app.jinja_env.globals["tr_text"] = t35_text
    app.jinja_env.globals["month_name_i18n"] = month_name_i18n_v35
    app.jinja_env.globals["status_i18n"] = status_i18n_v35
    app.jinja_env.globals["leave_label_i18n"] = leave_label_i18n_v35
except Exception:
    pass

with app.app_context():
    seed_db()
    auto_backup_once_per_day()

if __name__ == "__main__":
    os.environ["OAUTHLIB_INSECURE_TRANSPORT"]="1"
    app.run(debug=False)


@app.route("/setup_initial_balance", methods=["GET", "POST"])
@login_required
@subscription_required
def setup_initial_balance():
    u = current_user()
    if request.method == "POST":
        try:
            u.hire_date = datetime.strptime(request.form.get("hire_date"), "%Y-%m-%d").date()
            u.initial_balance = float(request.form.get("initial_balance") or 0)
            raw_date = request.form.get("initial_balance_date")
            u.initial_balance_date = datetime.strptime(raw_date, "%Y-%m-%d").date() if raw_date else date.today()
            db.session.commit()
            flash(t35("save") if "t35" in globals() else "Saved", "success")
            return redirect(url_for("dashboard"))
        except Exception as e:
            db.session.rollback()
            flash(str(e), "danger")
    return render_template("setup_initial_balance.html")
